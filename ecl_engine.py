"""ECL Computation Engine — refactored from ECL_Automation_v1.py"""

import re
import math
import numpy as np
import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Pure-Python replacements for scipy.stats.norm ────────────────────────────
# Eliminates the heavy scipy dependency (30MB+ with Fortran compilation).

class _Norm:
    """Drop-in replacement for scipy.stats.norm with cdf() and ppf()."""

    @staticmethod
    def cdf(x):
        """Standard normal CDF using Abramowitz & Stegun approximation."""
        a1, a2, a3 = 0.254829592, -0.284496736, 1.421413741
        a4, a5, p  = -1.453152027, 1.061405429, 0.3275911
        sign = -1 if x < 0 else 1
        ax = abs(x) / math.sqrt(2)
        t = 1.0 / (1.0 + p * ax)
        y = 1.0 - (((((a5*t + a4)*t) + a3)*t + a2)*t + a1) * t * math.exp(-ax*ax)
        return 0.5 * (1.0 + sign * y)

    @staticmethod
    def ppf(p):
        """Standard normal inverse CDF (probit) — rational approximation."""
        if p <= 0:
            return float('-inf')
        if p >= 1:
            return float('inf')
        if p == 0.5:
            return 0.0
        a = [-3.969683028665376e1, 2.209460984245205e2, -2.759285104469687e2,
              1.383577518672690e2, -3.066479806614716e1,  2.506628277459239e0]
        b = [-5.447609879822406e1, 1.615858368580409e2, -1.556989798598866e2,
              6.680131188771972e1, -1.328068155288572e1]
        c = [-7.784894002430293e-3, -3.223964580411365e-1, -2.400758277161838e0,
             -2.549732539343734e0,   4.374664141464968e0,   2.938163982698783e0]
        d = [ 7.784695709041462e-3,  3.224671290700398e-1,  2.445134137142996e0,
              3.754408661907416e0]
        p_low, p_high = 0.02425, 1 - 0.02425
        if p < p_low:
            q = math.sqrt(-2 * math.log(p))
            return (((((c[0]*q+c[1])*q+c[2])*q+c[3])*q+c[4])*q+c[5]) / \
                   ((((d[0]*q+d[1])*q+d[2])*q+d[3])*q+1)
        elif p <= p_high:
            q = p - 0.5
            r = q * q
            return (((((a[0]*r+a[1])*r+a[2])*r+a[3])*r+a[4])*r+a[5])*q / \
                   (((((b[0]*r+b[1])*r+b[2])*r+b[3])*r+b[4])*r+1)
        else:
            q = math.sqrt(-2 * math.log(1 - p))
            return -(((((c[0]*q+c[1])*q+c[2])*q+c[3])*q+c[4])*q+c[5]) / \
                    ((((d[0]*q+d[1])*q+d[2])*q+d[3])*q+1)

norm = _Norm()

# ── Constants ────────────────────────────────────────────────────────────────

FROM_STATES  = ["0", "1", "2", "3", "4"]
TO_STATES    = ["0", "1", "2", "3", "4", "WO", "ARC", "Closed"]
FROM_BUCKETS = ["0", "1-30", "31-60", "61-90", "90+"]
TO_BUCKETS   = ["0", "1-30", "31-60", "61-90", "90+", "WO", "ARC", "Closed"]
DEFAULT_TO   = {"4", "WO", "ARC"}
DEFAULT_TO_B = {"90+", "WO", "ARC"}

MONTH_MAP = {
    "jan": 1, "feb": 2,  "mar": 3,  "apr": 4,
    "may": 5, "jun": 6,  "jul": 7,  "aug": 8,
    "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    "june": 6, "july": 7,
}
MONTH_LABELS = {
    1:"Jan", 2:"Feb",  3:"Mar",  4:"Apr",
    5:"May", 6:"Jun",  7:"Jul",  8:"Aug",
    9:"Sep", 10:"Oct", 11:"Nov", 12:"Dec",
}
DPD_BUCKET = {
    "0":"0", "1":"1-30", "2":"31-60", "3":"61-90", "4":"90+",
    "WO":"WO", "ARC":"ARC", "Closed":"Closed",
}

SERIES_MAP = {
    "LUR":         "Unemployment Rate",
    "NGDP_RPCH":   "GDP",
    "PCPIPCH":     "Inflation",
    "GGX_NGDP":    "Govt_Expenditure",
    "GGXWDG_NGDP": "General government gross debt",
}
SERIES_ORDER  = ["LUR", "NGDP_RPCH", "PCPIPCH", "GGX_NGDP", "GGXWDG_NGDP"]
GDP_CODE      = "NGDP_RPCH"
GDP_PRECISE   = {2025: 6.198, 2026: 6.265, 2027: 6.471}

# ── Styles ───────────────────────────────────────────────────────────────────

FILL_DARK   = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
FILL_MED    = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
FILL_LIGHT  = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
FILL_ALT    = PatternFill("solid", start_color="EBF5FB", end_color="EBF5FB")
FILL_WHITE  = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
FILL_GREEN  = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
FILL_WARN   = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
FILL_RED    = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
FILL_GREY   = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
FILL_GOLD   = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
FILL_HIST   = PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1")
FILL_FCST   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
FILL_EXTRAP = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
FILL_BASE   = PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1")
FILL_UP     = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
FILL_DOWN   = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
FILL_PARAM  = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")

SCEN_HDR_FILL = {
    "Base":     PatternFill("solid", start_color="2E75B6", end_color="2E75B6"),
    "Upturn":   PatternFill("solid", start_color="375623", end_color="375623"),
    "Downturn": PatternFill("solid", start_color="C00000", end_color="C00000"),
}
SCEN_ROW_FILL = {"Base": FILL_BASE, "Upturn": FILL_UP, "Downturn": FILL_DOWN}

FONT_TITLE   = Font(name="Arial", bold=True,  color="1F4E79", size=12)
FONT_HDR     = Font(name="Arial", bold=True,  color="FFFFFF", size=9)
FONT_LABEL   = Font(name="Arial", bold=True,  color="1F4E79", size=9)
FONT_BODY    = Font(name="Arial", size=9)
FONT_TOTAL   = Font(name="Arial", bold=True,  size=9)
FONT_PARAM   = Font(name="Arial", bold=True,  color="7F4F00", size=9)
FONT_SECTION = Font(name="Arial", bold=True,  color="FFFFFF", size=10)
FONT_ODR     = Font(name="Arial", bold=True,  color="1F4E79", size=9)
FONT_GREY    = Font(name="Arial", italic=True, color="808080", size=9)
FONT_NA      = Font(name="Arial", bold=True,  color="C00000", size=9)
FONT_NOTE    = Font(name="Arial", italic=True, color="595959", size=8)
FONT_Z_POS   = Font(name="Arial", size=9, color="375623")
FONT_Z_NEG   = Font(name="Arial", size=9, color="C00000")
FONT_BASE    = Font(name="Arial", size=9, color="1F4E79")
FONT_UP      = Font(name="Arial", size=9, color="375623")
FONT_DOWN    = Font(name="Arial", size=9, color="C00000")
SCEN_Z_FONT  = {"Base": FONT_BASE, "Upturn": FONT_UP, "Downturn": FONT_DOWN}

_thin  = Side(style="thin", color="B0C4DE")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
ALIGN_C = Alignment(horizontal="center", vertical="center")
ALIGN_R = Alignment(horizontal="right",  vertical="center")
ALIGN_L = Alignment(horizontal="left",   vertical="center")

FMT_INT  = "#,##0"
FMT_PCT2 = "0.00%"
FMT_PCT6 = "0.000000%"
FMT_NUM2 = "0.00"
FMT_NUM6 = "0.000000"
FMT_DATE = "YYYY-MM-DD"


# ── Helpers ──────────────────────────────────────────────────────────────────

def sc(cell, fill=None, font=None, align=None, fmt=None, border=None, val=None):
    if val is not None:  cell.value         = val
    if fill:             cell.fill          = fill
    if font:             cell.font          = font
    if align:            cell.alignment     = align
    if fmt:              cell.number_format = fmt
    if border:           cell.border        = border


def parse_dpd_col(col):
    m = re.match(r"DPD_([A-Za-z]+)-(\d{2})$", col)
    if not m:
        return None
    month = MONTH_MAP.get(m.group(1).lower())
    return (month, 2000 + int(m.group(2))) if month else None


def cast_dpd(v):
    if pd.isna(v):
        return np.nan
    try:
        return str(int(float(v)))
    except (ValueError, TypeError):
        return str(v)


def basel_retail_rho(pd_val):
    if pd_val >= 1.0:
        return 0.03
    w = (1 - np.exp(-35 * pd_val)) / (1 - np.exp(-35))
    return 0.03 * w + 0.16 * (1 - w)


def vasicek_pd(ttc, rho, z):
    if ttc >= 1.0:
        return 1.0
    return norm.cdf((norm.ppf(ttc) - np.sqrt(rho) * z) / np.sqrt(1 - rho))


def year_fill(yr, hist_cutoff):
    if yr <= hist_cutoff: return FILL_HIST
    if yr <= 2027:        return FILL_FCST
    return FILL_EXTRAP


def year_font(yr, hist_cutoff):
    if yr <= hist_cutoff: return Font(name="Arial", size=9, color="1F4E79")
    if yr <= 2027:        return Font(name="Arial", size=9, color="375623")
    return Font(name="Arial", size=9, color="7F4F00", italic=True)


# ── Engine ───────────────────────────────────────────────────────────────────

class ECLEngine:
    def __init__(self, dpd_path, weo_path, output_path, config=None):
        self.dpd_path    = dpd_path
        self.weo_path    = weo_path
        self.output_path = output_path

        cfg = config or {}
        self.tm_start_year = cfg.get("tm_start_year", 2020)
        self.shock         = cfg.get("shock", 0.10)
        self.hist_cutoff   = cfg.get("hist_cutoff", 2024)
        self.odr_display_years = cfg.get("odr_display_years", list(range(2020, 2025)))
        self.scen_weights  = cfg.get("scen_weights", {"Base": 0.50, "Upturn": 0.05, "Downturn": 0.45})

        self.weo_years     = list(range(2019, 2028))
        self.extrap_years  = list(range(2028, 2033))
        self.trend_window  = [2025, 2026, 2027]
        self.display_years = self.weo_years + self.extrap_years
        self.calib_years   = list(range(2019, 2026))
        self.forecast_yrs  = list(range(2025, 2033))
        self.scenarios     = ["Base", "Upturn", "Downturn"]

    # ── Step 1: Load DPD ─────────────────────────────────────────────────

    def _load_dpd(self):
        self.df_dpd = pd.read_excel(self.dpd_path)
        self.df_dpd.columns = self.df_dpd.columns.str.strip()

        self.col_lookup = {
            parse_dpd_col(c): c for c in self.df_dpd.columns if parse_dpd_col(c)
        }

        self.tm_pairs = {}
        for month in range(1, 13):
            pairs, yr = [], self.tm_start_year
            while (month, yr) in self.col_lookup and (month, yr + 1) in self.col_lookup:
                pairs.append((yr, yr + 1))
                yr += 1
            self.tm_pairs[month] = pairs

        self.pairs_by_yr = defaultdict(list)
        for (m, yr) in self.col_lookup:
            if (m, yr + 1) in self.col_lookup:
                self.pairs_by_yr[yr].append(m)
        self.year_pairs = sorted(self.pairs_by_yr.keys())

    # ── Step 2: Transition matrix ────────────────────────────────────────

    def _compute_transition(self, from_col, to_col):
        frm = self.df_dpd[from_col].map(cast_dpd)
        to_ = self.df_dpd[to_col].map(cast_dpd)
        rows = {}
        for fs in FROM_STATES:
            mask   = frm == fs
            subset = to_[mask]
            total  = int(mask.sum())
            counts = {ts: int((subset == ts).sum()) for ts in TO_STATES}
            dft    = sum(counts[d] for d in DEFAULT_TO)
            rows[fs] = {**counts, "Total": total,
                        "Default": dft / total if total else 0.0}
        agg   = {ts: sum(rows[fs][ts] for fs in FROM_STATES) for ts in TO_STATES}
        grand = sum(rows[fs]["Total"] for fs in FROM_STATES)
        agg["Total"]   = grand
        agg["Default"] = sum(agg[d] for d in DEFAULT_TO) / grand if grand else 0.0
        rows["Total"]  = agg
        return rows

    # ── Step 3: ODR ──────────────────────────────────────────────────────

    def _compute_odr_matrix(self, from_yr):
        agg         = {fb: {tb: 0 for tb in TO_BUCKETS} for fb in FROM_BUCKETS}
        months_used = 0
        total_obs   = 0

        for m in sorted(self.pairs_by_yr[from_yr]):
            frm_s = self.df_dpd[self.col_lookup[(m, from_yr)]].map(cast_dpd).map(
                        lambda v: DPD_BUCKET.get(v, np.nan))
            to_s  = self.df_dpd[self.col_lookup[(m, from_yr + 1)]].map(cast_dpd).map(
                        lambda v: DPD_BUCKET.get(v, np.nan))
            obs = int(frm_s.isin(FROM_BUCKETS).sum())
            if obs == 0:
                continue
            months_used += 1
            total_obs   += obs
            for fb in FROM_BUCKETS:
                mask = frm_s == fb
                sub  = to_s[mask]
                for tb in TO_BUCKETS:
                    agg[fb][tb] += int((sub == tb).sum())
        return agg, months_used, total_obs

    def _compute_odr(self):
        self.odr_results = {}
        for from_yr in self.year_pairs:
            mat, nm, total_obs = self._compute_odr_matrix(from_yr)
            tot   = {tb: sum(mat[fb][tb] for fb in FROM_BUCKETS) for tb in TO_BUCKETS}
            grand = sum(tot.values())
            odr   = sum(tot[d] for d in DEFAULT_TO_B) / grand if grand else None
            self.odr_results[from_yr] = {
                "matrix": mat, "odr": odr,
                "months": nm,  "total_obs": total_obs,
            }

    # ── Step 4: TTC & Rho ───────────────────────────────────────────────

    def _compute_ttc_rho(self):
        ttc_n_years = len(self.odr_display_years)
        odr_by_grade = {fb: {} for fb in FROM_BUCKETS}
        for yr in self.odr_display_years:
            if yr not in self.pairs_by_yr:
                continue
            mat, nm, _ = self._compute_odr_matrix(yr)
            if nm == 0:
                continue
            for fb in FROM_BUCKETS:
                row   = mat[fb]
                total = sum(row.values())
                if total == 0:
                    odr_by_grade[fb][yr] = 0.0
                    continue
                dft = sum(row[d] for d in DEFAULT_TO_B)
                odr_by_grade[fb][yr] = dft / total

        self.ttc = {
            fb: sum(odr_by_grade[fb].values()) / ttc_n_years
            for fb in FROM_BUCKETS
        }
        self.ttc["90+"] = 1.0
        self.rho = {fb: basel_retail_rho(self.ttc[fb]) for fb in FROM_BUCKETS}

    # ── Step 5: WEO ─────────────────────────────────────────────────────

    def _load_weo(self):
        df_weo = pd.read_excel(self.weo_path, header=None)

        weo_year_cols = {}
        for c, v in enumerate(df_weo.iloc[0]):
            try:
                yr = int(float(v))
                if 2000 <= yr <= 2027:
                    weo_year_cols[yr] = c
            except (ValueError, TypeError):
                pass

        self.raw_weo = {}
        for _, row in df_weo.iterrows():
            code = str(row[1]).strip()
            if code in SERIES_MAP:
                self.raw_weo[code] = {yr: float(row[weo_year_cols[yr]]) for yr in self.weo_years}

    def _linear_extrap(self, series, horizon):
        base = self.trend_window[0]
        x    = np.array([yr - base for yr in self.trend_window], dtype=float)
        y    = np.array([series[yr] for yr in self.trend_window], dtype=float)
        slope, intercept = np.polyfit(x, y, 1)
        return round(intercept + slope * (horizon - base), 4)

    def _compute_mav(self):
        self.mav = {}
        for code in SERIES_ORDER:
            self.mav[code] = {}
            for yr in self.weo_years:
                self.mav[code][yr] = round(self.raw_weo[code][yr], 4)
            for yr in self.extrap_years:
                self.mav[code][yr] = self._linear_extrap(self.raw_weo[code], yr)

        self.mav_params = {}
        for code in SERIES_ORDER:
            vals = np.array([self.mav[code][yr] for yr in self.calib_years])
            self.mav_params[code] = {
                "LTM": round(float(vals.mean()), 2),
                "SD":  round(float(vals.std(ddof=0)), 2),
            }

        self.z_factors = {}
        for code in SERIES_ORDER:
            ltm = self.mav_params[code]["LTM"]
            sd  = self.mav_params[code]["SD"]
            self.z_factors[code] = {
                yr: round((GDP_PRECISE.get(yr, self.mav[code][yr]) if code == GDP_CODE
                           else self.mav[code][yr]) / sd - ltm / sd, 2)
                for yr in self.display_years
            }

    # ── Step 6: Scenarios ────────────────────────────────────────────────

    def _compute_scenarios(self):
        self.gdp_ltm = self.mav_params[GDP_CODE]["LTM"]
        self.gdp_sd  = self.mav_params[GDP_CODE]["SD"]

        self.gdp_z_raw = {
            yr: (GDP_PRECISE.get(yr, self.mav[GDP_CODE][yr]) - self.gdp_ltm) / self.gdp_sd
            for yr in self.display_years
        }
        self.mev_scenarios = {
            yr: {
                "Base":     round(z, 2),
                "Upturn":   round(z + abs(z) * self.shock, 2),
                "Downturn": round(z - abs(z) * self.shock, 2),
            }
            for yr, z in self.gdp_z_raw.items()
        }
        self.scen_z_raw = {
            yr: {
                "Base":     self.gdp_z_raw[yr],
                "Upturn":   self.gdp_z_raw[yr] + abs(self.gdp_z_raw[yr]) * self.shock,
                "Downturn": self.gdp_z_raw[yr] - abs(self.gdp_z_raw[yr]) * self.shock,
            }
            for yr in self.forecast_yrs
        }

    # ── Step 7: Vasicek PD ──────────────────────────────────────────────

    def _compute_vasicek(self):
        self.pd_results = {}
        for scen in self.scenarios:
            self.pd_results[scen] = {
                grade: {
                    yr: vasicek_pd(self.ttc[grade], self.rho[grade],
                                   self.scen_z_raw[yr][scen])
                    for yr in self.forecast_yrs
                }
                for grade in FROM_BUCKETS
            }

    # ── Step 8: Survival Analysis ────────────────────────────────────────

    def _compute_survival(self):
        self.surv_1 = {}
        self.cumul_surv = {}
        for scen in self.scenarios:
            self.surv_1[scen] = {}
            self.cumul_surv[scen] = {}
            for grade in FROM_BUCKETS:
                self.surv_1[scen][grade] = {}
                self.cumul_surv[scen][grade] = {}
                running = 1.0
                for yr in self.forecast_yrs:
                    sp = 1.0 - self.pd_results[scen][grade][yr]
                    self.surv_1[scen][grade][yr] = sp
                    running *= sp
                    self.cumul_surv[scen][grade][yr] = running

    # ── Step 9: PIT PD ──────────────────────────────────────────────────

    def _compute_pit_pd(self):
        self.marginal_pd = {}
        for scen in self.scenarios:
            self.marginal_pd[scen] = {}
            for grade in FROM_BUCKETS:
                self.marginal_pd[scen][grade] = {}
                for i, yr in enumerate(self.forecast_yrs):
                    if i == 0:
                        self.marginal_pd[scen][grade][yr] = self.pd_results[scen][grade][yr]
                    else:
                        prev_yr = self.forecast_yrs[i - 1]
                        self.marginal_pd[scen][grade][yr] = (
                            self.cumul_surv[scen][grade][prev_yr]
                            - self.cumul_surv[scen][grade][yr]
                        )

        self.pit_pd_vals = {}
        for grade in FROM_BUCKETS:
            self.pit_pd_vals[grade] = {}
            for yr in self.forecast_yrs:
                self.pit_pd_vals[grade][yr] = sum(
                    self.scen_weights[s] * self.marginal_pd[s][grade][yr]
                    for s in self.scenarios
                )

        last_yr = self.forecast_yrs[-1]
        self.lifetime_pd = {
            grade: 1.0 - sum(
                self.scen_weights[s] * self.cumul_surv[s][grade][last_yr]
                for s in self.scenarios
            )
            for grade in FROM_BUCKETS
        }

    # ── Excel generation ─────────────────────────────────────────────────

    def _generate_excel(self):
        wb = Workbook()
        self._sheet_tm(wb)
        self._sheet_odr(wb)
        self._sheet_mav(wb)
        self._sheets_mev_detail(wb)
        self._sheet_mev_scenarios(wb)
        self._sheet_vasicek(wb)
        self._sheet_pd_comparison(wb)
        self._sheet_inputs(wb)
        self._sheet_survival(wb)
        self._sheet_pit_pd(wb)
        wb.save(self.output_path)

    def _sheet_tm(self, wb):
        ws = wb.active
        ws.title = "Transition Matrix"
        ws.sheet_view.showGridLines = False
        TM_COLS = 12
        TM_ROWS = 10
        TO_KEYS    = ["0","1","2","3","4","WO","ARC","Closed","Total","Default"]
        COL_LABELS = ["0","1","2.0","3.0","4.0","WO","ARC","Closed","Total","Defualt"]
        max_years = max((len(p) for p in self.tm_pairs.values()), default=0)
        for yi in range(max_years):
            bc = yi * TM_COLS + 1
            ws.column_dimensions[get_column_letter(bc)].width      = 2
            ws.column_dimensions[get_column_letter(bc + 1)].width  = 8
            ws.column_dimensions[get_column_letter(bc + 2)].width  = 12
            ws.column_dimensions[get_column_letter(bc + 3)].width  = 8
            for dc in range(4, 10):
                ws.column_dimensions[get_column_letter(bc + dc)].width = 9
            ws.column_dimensions[get_column_letter(bc + 10)].width = 10
            ws.column_dimensions[get_column_letter(bc + 11)].width = 11

        for month in range(1, 13):
            pairs = self.tm_pairs[month]
            if not pairs:
                continue
            base_row = (month - 1) * TM_ROWS + 1
            ws.row_dimensions[base_row].height     = 15
            ws.row_dimensions[base_row + 1].height = 15
            for ri in range(6):
                ws.row_dimensions[base_row + 2 + ri].height = 14
            for yi, (from_yr, to_yr) in enumerate(pairs):
                bc  = yi * TM_COLS + 1
                c1  = bc + 1
                from_date = pd.Timestamp(f"{from_yr}-{month:02d}-01")
                to_date   = pd.Timestamp(f"{to_yr}-{month:02d}-01")
                sc(ws.cell(base_row, c1),     fill=FILL_DARK, font=FONT_HDR,
                   align=ALIGN_C, border=BORDER, val=f"TM{month}")
                sc(ws.cell(base_row, c1+1),   fill=FILL_DARK, font=FONT_HDR,
                   align=ALIGN_C, fmt=FMT_DATE, border=BORDER, val=to_date)
                sc(ws.cell(base_row, c1+2),   fill=FILL_DARK, font=FONT_HDR,
                   align=ALIGN_C, border=BORDER, val=f"Year {yi+1}")
                for dc in range(3, 11):
                    sc(ws.cell(base_row, c1+dc), fill=FILL_DARK, border=BORDER)
                sc(ws.cell(base_row+1, c1), fill=FILL_MED, font=FONT_HDR,
                   align=ALIGN_C, fmt=FMT_DATE, border=BORDER, val=from_date)
                for di, lbl in enumerate(COL_LABELS):
                    sc(ws.cell(base_row+1, c1+1+di), fill=FILL_MED, font=FONT_HDR,
                       align=ALIGN_C, border=BORDER, val=lbl)
                mat = self._compute_transition(
                    self.col_lookup[(month, from_yr)],
                    self.col_lookup[(month, to_yr)])
                for ri, fs in enumerate(FROM_STATES + ["Total"]):
                    r        = base_row + 2 + ri
                    is_total = fs == "Total"
                    row_fill = FILL_LIGHT if is_total else (FILL_ALT if ri % 2 == 0 else FILL_WHITE)
                    row_font = FONT_TOTAL if is_total else FONT_BODY
                    sc(ws.cell(r, c1), fill=row_fill, font=row_font,
                       align=ALIGN_C, border=BORDER, val="Total" if is_total else fs)
                    for di, key in enumerate(TO_KEYS):
                        val = mat[fs][key]
                        fmt = FMT_PCT2 if key == "Default" else FMT_INT
                        sc(ws.cell(r, c1+1+di), fill=row_fill, font=row_font,
                           align=ALIGN_R, fmt=fmt, border=BORDER, val=val)
        ws.freeze_panes = "B3"

    def _sheet_odr(self, wb):
        ws = wb.create_sheet("ODR")
        ws.sheet_view.showGridLines = False

        # Filter to display years only
        odr_yrs = [yr for yr in self.year_pairs if yr in self.odr_display_years]
        num_odr_yrs = len(odr_yrs)
        to_cols_odr = list(TO_BUCKETS)
        cols_per_yr = len(to_cols_odr) + 2  # 8 to-buckets + Total + Default = 10
        sub_labels = to_cols_odr + ["Total", "Defualt"]

        # Row 1: Title
        ws.row_dimensions[1].height = 22
        sc(ws.cell(1, 1), val="Observed Default Rate",
           font=FONT_TITLE, align=ALIGN_L)
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=1 + num_odr_yrs * cols_per_yr)
        ws.row_dimensions[2].height = 6

        # Row 3: Year-group headers
        ws.row_dimensions[3].height = 16
        sc(ws.cell(3, 1), fill=FILL_DARK, border=BORDER)

        for yi, from_yr in enumerate(odr_yrs):
            bc = 2 + yi * cols_per_yr
            res = self.odr_results[from_yr]
            partial = 0 < res["months"] < 12
            no_data = res["months"] == 0

            if no_data:
                hdr_fill = FILL_GREY
                hdr_font = Font(name="Arial", bold=True, color="808080", size=9)
                lbl = f"Year {yi+1}  ({from_yr}\u2192{from_yr+1})  [No data]"
            elif partial:
                hdr_fill = FILL_WARN
                hdr_font = Font(name="Arial", bold=True, color="7F4F00", size=9)
                lbl = f"Year {yi+1}  ({from_yr}\u2192{from_yr+1})  [{res['months']} months]*"
            else:
                hdr_fill, hdr_font = FILL_DARK, FONT_HDR
                lbl = f"Year {yi+1}  ({from_yr}\u2192{from_yr+1})"

            sc(ws.cell(3, bc), val=lbl,
               fill=hdr_fill, font=hdr_font, align=ALIGN_C, border=BORDER)
            ws.merge_cells(start_row=3, start_column=bc,
                           end_row=3, end_column=bc + cols_per_yr - 1)

        # Row 4: Sub-column headers
        ws.row_dimensions[4].height = 15
        sc(ws.cell(4, 1), val="From \\ To",
           fill=FILL_MED, font=FONT_HDR, align=ALIGN_C, border=BORDER)

        for yi in range(num_odr_yrs):
            bc = 2 + yi * cols_per_yr
            for di, lbl in enumerate(sub_labels):
                sc(ws.cell(4, bc + di), val=lbl,
                   fill=FILL_MED, font=FONT_HDR, align=ALIGN_C, border=BORDER)

        # Rows 5-10: Data rows
        for ri, fb in enumerate(FROM_BUCKETS + ["Total"]):
            row_e = 5 + ri
            is_total = fb == "Total"
            ws.row_dimensions[row_e].height = 14

            row_fill = FILL_LIGHT if is_total else (FILL_ALT if ri % 2 == 0 else FILL_WHITE)
            row_font = FONT_TOTAL if is_total else FONT_BODY

            sc(ws.cell(row_e, 1), val=fb,
               fill=row_fill,
               font=Font(name="Arial", bold=True, size=9, color="1F4E79"),
               align=ALIGN_C, border=BORDER)

            for yi, from_yr in enumerate(odr_yrs):
                bc = 2 + yi * cols_per_yr
                res = self.odr_results[from_yr]
                no_data = res["months"] == 0
                mat = res["matrix"]

                if no_data:
                    for di in range(cols_per_yr):
                        sc(ws.cell(row_e, bc + di), val="N/A",
                           fill=FILL_GREY, font=FONT_GREY, align=ALIGN_C, border=BORDER)
                    continue

                if is_total:
                    row_data = {tb: sum(mat[b][tb] for b in FROM_BUCKETS) for tb in to_cols_odr}
                else:
                    row_data = {tb: mat[fb][tb] for tb in to_cols_odr}

                total_val = sum(row_data.values())
                dft_val = (sum(row_data[d] for d in DEFAULT_TO_B if d in row_data) / total_val
                           if total_val else 0.0)

                for di, tk in enumerate(to_cols_odr + ["Total", "Default"]):
                    col_e = bc + di
                    if tk == "Total":
                        v, fmt = total_val, FMT_INT
                    elif tk == "Default":
                        v, fmt = dft_val, FMT_PCT2
                    else:
                        v, fmt = row_data[tk], FMT_INT

                    if not is_total and tk in DEFAULT_TO_B:
                        c_fill = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
                        c_font = Font(name="Arial", bold=True, color="C00000", size=9)
                    elif not is_total and tk == "Default":
                        c_fill = FILL_RED
                        c_font = Font(name="Arial", bold=True, color="C00000", size=9)
                    else:
                        c_fill, c_font = row_fill, row_font

                    sc(ws.cell(row_e, col_e), val=v,
                       fill=c_fill, font=c_font,
                       align=ALIGN_R, fmt=fmt, border=BORDER)

        # Spacers
        ws.row_dimensions[11].height = 8
        ws.row_dimensions[12].height = 8

        # Rows 13-14: ODR Summary
        ws.row_dimensions[13].height = 15
        sc(ws.cell(13, 1), val="Year",
           fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)

        ws.row_dimensions[14].height = 15
        sc(ws.cell(14, 1), val="ODR",
           fill=FILL_GREEN, font=FONT_ODR, align=ALIGN_C, border=BORDER)

        for yi, from_yr in enumerate(odr_yrs):
            col_e = 2 + yi
            res = self.odr_results[from_yr]
            no_data = res["months"] == 0
            partial = 0 < res["months"] < 12
            fill = FILL_RED if no_data else (FILL_WARN if partial else FILL_GREEN)

            sc(ws.cell(13, col_e), val=f"Year {yi+1}",
               fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)

            if no_data:
                sc(ws.cell(14, col_e), val="N/A",
                   fill=fill, font=FONT_NA, align=ALIGN_C, border=BORDER)
            else:
                sc(ws.cell(14, col_e), val=res["odr"],
                   fill=fill, font=FONT_ODR, align=ALIGN_R, fmt=FMT_PCT2, border=BORDER)

        # Notes
        for row_n, (text, colour, italic) in enumerate([
            ("Notes:", "1F4E79", False),
            ("  Columns 90+, WO, ARC are default states (highlighted red).  "
             "Amber (*) = Partial year - ODR may be understated.", "7F4F00", True),
            ("  ODR = Total defaults (90+, WO, ARC) / Total observations across all from-buckets.",
             "595959", True),
        ], start=15):
            ws.row_dimensions[row_n].height = 12
            sc(ws.cell(row_n, 1), val=text,
               font=Font(name="Arial", size=8, italic=italic, color=colour), align=ALIGN_L)
            ws.merge_cells(start_row=row_n, start_column=1,
                           end_row=row_n, end_column=1 + num_odr_yrs * cols_per_yr)

        # Column widths
        ws.column_dimensions["A"].width = 11
        sub_col_widths = [6, 7, 7, 7, 7, 6, 6, 8, 8, 9]
        for yi in range(num_odr_yrs):
            bc = 2 + yi * cols_per_yr
            for di, w in enumerate(sub_col_widths):
                ws.column_dimensions[get_column_letter(bc + di)].width = w

    def _sheet_mav(self, wb):
        ws = wb.create_sheet("MAV Summary")
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 22
        sc(ws.cell(1, 1),
           val="MAV Index - Z-Factors Summary (All MEVs, India, IMF WEO Apr 2025)",
           font=FONT_TITLE, align=ALIGN_L)
        ws.merge_cells(start_row=1, start_column=1, end_row=1,
                       end_column=5+len(self.display_years))
        ws.row_dimensions[2].height = 8
        ws.row_dimensions[3].height = 16
        for ci, lbl in enumerate(["MEV", "LTM", "SD", "CV"]):
            sc(ws.cell(3, 1+ci), val=lbl,
               fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
        for ci, yr in enumerate(self.display_years):
            hf = (FILL_DARK if yr <= self.hist_cutoff else
                  FILL_MED  if yr <= 2027 else
                  PatternFill("solid", start_color="7F6000", end_color="7F6000"))
            sc(ws.cell(3, 5+ci), val=str(yr),
               fill=hf, font=FONT_HDR, align=ALIGN_C, border=BORDER)
        for ri, code in enumerate(SERIES_ORDER):
            row_e    = 4 + ri
            row_fill = FILL_ALT if ri % 2 == 0 else FILL_WHITE
            ws.row_dimensions[row_e].height = 15
            ltm_val = self.mav_params[code]["LTM"]
            sd_val  = self.mav_params[code]["SD"]
            cv_val  = abs(sd_val / ltm_val) if ltm_val != 0 else None
            sc(ws.cell(row_e, 1), val=SERIES_MAP[code],
               fill=row_fill, font=FONT_LABEL, align=ALIGN_L, border=BORDER)
            sc(ws.cell(row_e, 2), val=ltm_val,
               fill=FILL_GOLD, font=FONT_PARAM, align=ALIGN_R, fmt=FMT_NUM2, border=BORDER)
            sc(ws.cell(row_e, 3), val=sd_val,
               fill=FILL_GOLD, font=FONT_PARAM, align=ALIGN_R, fmt=FMT_NUM2, border=BORDER)
            sc(ws.cell(row_e, 4), val=cv_val,
               fill=FILL_GOLD, font=FONT_PARAM, align=ALIGN_R, fmt="0.00%", border=BORDER)
            for ci, yr in enumerate(self.display_years):
                z = self.z_factors[code][yr]
                sc(ws.cell(row_e, 5+ci), val=z,
                   fill=year_fill(yr, self.hist_cutoff),
                   font=FONT_Z_NEG if z < 0 else FONT_Z_POS,
                   align=ALIGN_R, fmt=FMT_NUM2, border=BORDER)
        ws.row_dimensions[9].height = 8
        sc(ws.cell(10, 1),
           val=(f"Z = (Value - LTM) / SD  |  CV = |SD / LTM|  |  Window {self.calib_years[0]}-{self.calib_years[-1]} "
                f"(n={len(self.calib_years)}, pop. std)  |  "
                "Blue=Actual  Green=IMF Forecast  Amber=Extrapolated"),
           font=FONT_NOTE, align=ALIGN_L)
        ws.merge_cells(start_row=10, start_column=1, end_row=10,
                       end_column=5+len(self.display_years))
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 9
        for ci in range(len(self.display_years)):
            ws.column_dimensions[get_column_letter(5+ci)].width = 6.5

    def _sheets_mev_detail(self, wb):
        for mi, code in enumerate(SERIES_ORDER):
            mev_name = SERIES_MAP[code]
            ws = wb.create_sheet(mev_name[:31])
            ws.sheet_view.showGridLines = False
            ltm = self.mav_params[code]["LTM"]
            sd  = self.mav_params[code]["SD"]
            zf  = self.z_factors[code]
            ws.row_dimensions[1].height = 18
            sc(ws.cell(1, 2), val=mi+1,
               font=Font(name="Arial", bold=True, size=11, color="1F4E79"), align=ALIGN_C)
            ws.row_dimensions[2].height = 18
            sc(ws.cell(2, 1), val="Selected MEV", font=FONT_LABEL, align=ALIGN_L)
            sc(ws.cell(2, 2), val=mev_name,
               font=Font(name="Arial", bold=True, size=10, color="1F4E79"), align=ALIGN_L)
            ws.row_dimensions[3].height = 8
            ws.row_dimensions[4].height = 16
            sc(ws.cell(4, 1), val="Variable",
               fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
            for ci, yr in enumerate(self.display_years):
                sc(ws.cell(4, 2+ci), val=str(yr),
                   fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
            ws.row_dimensions[5].height = 15
            sc(ws.cell(5, 1), val=mev_name,
               fill=FILL_LIGHT, font=FONT_LABEL, align=ALIGN_L, border=BORDER)
            for ci, yr in enumerate(self.display_years):
                sc(ws.cell(5, 2+ci), val=round(self.mav[code][yr], 2),
                   fill=year_fill(yr, self.hist_cutoff), font=year_font(yr, self.hist_cutoff),
                   align=ALIGN_R, fmt=FMT_NUM2, border=BORDER)
            ws.row_dimensions[6].height = 8
            ws.row_dimensions[7].height = 8
            ws.row_dimensions[8].height = 16
            sc(ws.cell(8, 1), val="Scenario Development",
               fill=FILL_MED, font=FONT_SECTION, align=ALIGN_L, border=BORDER)
            ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=4)
            ws.row_dimensions[9].height = 15
            sc(ws.cell(9, 1), fill=FILL_MED, border=BORDER)
            sc(ws.cell(9, 2), val=mev_name,
               fill=FILL_MED, font=FONT_HDR, align=ALIGN_C, border=BORDER)
            ws.row_dimensions[10].height = 15
            sc(ws.cell(10, 1), val="Long Term Mean",
               fill=FILL_GOLD, font=FONT_PARAM, align=ALIGN_L, border=BORDER)
            sc(ws.cell(10, 2), val=ltm,
               fill=FILL_GOLD, font=FONT_PARAM, align=ALIGN_R, fmt=FMT_NUM2, border=BORDER)
            ws.row_dimensions[11].height = 15
            sc(ws.cell(11, 1), val="Standard Deviation",
               fill=FILL_GOLD, font=FONT_PARAM, align=ALIGN_L, border=BORDER)
            sc(ws.cell(11, 2), val=sd,
               fill=FILL_GOLD, font=FONT_PARAM, align=ALIGN_R, fmt=FMT_NUM2, border=BORDER)
            ws.row_dimensions[12].height = 8
            ws.row_dimensions[13].height = 8
            ws.row_dimensions[14].height = 16
            sc(ws.cell(14, 1), val="Z_Factor",
               fill=FILL_MED, font=FONT_SECTION, align=ALIGN_L, border=BORDER)
            ws.merge_cells(start_row=14, start_column=1, end_row=14, end_column=4)
            ws.row_dimensions[15].height = 15
            sc(ws.cell(15, 1), val="Variable",
               fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
            sc(ws.cell(15, 2), val="Relationship",
               fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
            for ci, yr in enumerate(self.display_years):
                sc(ws.cell(15, 3+ci), val=str(yr),
                   fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
            ws.row_dimensions[16].height = 15
            sc(ws.cell(16, 1), val=mev_name,
               fill=FILL_LIGHT, font=FONT_LABEL, align=ALIGN_L, border=BORDER)
            sc(ws.cell(16, 2), fill=FILL_LIGHT, border=BORDER)
            for ci, yr in enumerate(self.display_years):
                z = zf[yr]
                sc(ws.cell(16, 3+ci), val=z,
                   fill=year_fill(yr, self.hist_cutoff),
                   font=FONT_Z_NEG if z < 0 else FONT_Z_POS,
                   align=ALIGN_R, fmt=FMT_NUM2, border=BORDER)
            ws.row_dimensions[17].height = 8
            ws.row_dimensions[18].height = 13
            sc(ws.cell(18, 1),
               val=(f"Z = (Value - LTM) / SD  |  LTM={ltm}, SD={sd}  |  "
                    f"Window {self.calib_years[0]}-{self.calib_years[-1]} "
                    f"(n={len(self.calib_years)}, pop. std)  |  "
                    "Blue=Actual  Green=IMF Forecast  Amber=Extrapolated"),
               font=FONT_NOTE, align=ALIGN_L)
            ws.merge_cells(start_row=18, start_column=1,
                           end_row=18, end_column=2+len(self.display_years))
            ws.column_dimensions["A"].width = 34
            ws.column_dimensions["B"].width = 14
            for ci in range(len(self.display_years)):
                ws.column_dimensions[get_column_letter(3+ci)].width = 6.5

    def _sheet_mev_scenarios(self, wb):
        ws = wb.create_sheet("MEV Scenarios")
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 22
        sc(ws.cell(1, 1),
           val=(f"MEV Scenario - All Variables  "
                f"[GDP-driven  |  LTM={self.gdp_ltm}  |  SD={self.gdp_sd}  |  "
                f"Shock=+/-{int(self.shock*100)}%]"),
           font=FONT_TITLE, align=ALIGN_L)
        ws.merge_cells(start_row=1, start_column=1, end_row=1,
                       end_column=2+len(self.display_years))
        ws.row_dimensions[2].height = 8
        current_row = 3
        for mi, code in enumerate(SERIES_ORDER):
            mev_name = SERIES_MAP[code]
            ws.row_dimensions[current_row].height = 14
            sc(ws.cell(current_row, 1), val=f"  {mev_name}",
               fill=FILL_DARK, font=Font(name="Arial", bold=True, color="FFFFFF", size=9),
               align=ALIGN_L, border=BORDER)
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=2+len(self.display_years))
            current_row += 1
            for s in self.scenarios:
                ws.row_dimensions[current_row].height = 15
                sc(ws.cell(current_row, 1), val=s,
                   fill=SCEN_HDR_FILL[s],
                   font=Font(name="Arial", bold=True, color="FFFFFF", size=10),
                   align=ALIGN_L, border=BORDER)
                ws.merge_cells(start_row=current_row, start_column=1,
                               end_row=current_row, end_column=2+len(self.display_years))
                current_row += 1
                ws.row_dimensions[current_row].height = 14
                sc(ws.cell(current_row, 1), val="Variable",
                   fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
                for ci, yr in enumerate(self.display_years):
                    sc(ws.cell(current_row, 2+ci), val=str(yr),
                       fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
                current_row += 1
                ws.row_dimensions[current_row].height = 14
                sc(ws.cell(current_row, 1), val="MEV_Index",
                   fill=SCEN_ROW_FILL[s], font=FONT_LABEL, align=ALIGN_L, border=BORDER)
                for ci, yr in enumerate(self.display_years):
                    z = self.mev_scenarios[yr][s]
                    sc(ws.cell(current_row, 2+ci), val=z,
                       fill=year_fill(yr, self.hist_cutoff), font=SCEN_Z_FONT[s],
                       align=ALIGN_R, fmt=FMT_NUM2, border=BORDER)
                current_row += 1
                for _ in range(2):
                    ws.row_dimensions[current_row].height = 5
                    current_row += 1
            for _ in range(2):
                ws.row_dimensions[current_row].height = 6
                current_row += 1
        ws.row_dimensions[current_row].height = 13
        sc(ws.cell(current_row, 1),
           val=(f"Scenario driver: GDP Z-factor  |  "
                f"Base=Z  |  Upturn=Z+|Z|*{self.shock}  |  Downturn=Z-|Z|*{self.shock}  |  "
                "Blue=Actual  Green=IMF Forecast  Amber=Extrapolated"),
           font=FONT_NOTE, align=ALIGN_L)
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=2+len(self.display_years))
        ws.column_dimensions["A"].width = 16
        for ci in range(len(self.display_years)):
            ws.column_dimensions[get_column_letter(2+ci)].width = 6.5

    def _sheet_vasicek(self, wb):
        ws = wb.create_sheet("Vasicek PD")
        ws.sheet_view.showGridLines = False
        BLOCK_ROWS = 3 + len(FROM_BUCKETS) + 2
        ws.row_dimensions[1].height = 20
        sc(ws.cell(1, 1), val="Vasicek Unconditional Probability of Default",
           font=FONT_TITLE, align=ALIGN_L)
        ws.merge_cells(start_row=1, start_column=1, end_row=1,
                       end_column=4+len(self.forecast_yrs))
        ws.row_dimensions[2].height = 6
        for si, s in enumerate(self.scenarios):
            base_r   = 3 + si * BLOCK_ROWS
            row_fill = SCEN_ROW_FILL[s]
            ws.row_dimensions[base_r].height = 16
            sc(ws.cell(base_r, 1), val=s,
               fill=SCEN_HDR_FILL[s],
               font=Font(name="Arial", bold=True, color="FFFFFF", size=10),
               align=ALIGN_L, border=BORDER)
            ws.merge_cells(start_row=base_r, start_column=1,
                           end_row=base_r, end_column=4+len(self.forecast_yrs))
            ws.row_dimensions[base_r+1].height = 15
            for ci, hdr in enumerate(["Grades", "TTC", "Asset Correlation (\u03c1)"]):
                sc(ws.cell(base_r+1, 1+ci), val=hdr,
                   fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
            for ci, yr in enumerate(self.forecast_yrs):
                sc(ws.cell(base_r+1, 4+ci), val=str(yr),
                   fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
            ws.row_dimensions[base_r+2].height = 13
            sc(ws.cell(base_r+2, 1), val="Z-Factor ->",
               fill=FILL_PARAM, font=FONT_PARAM, align=ALIGN_R, border=BORDER)
            sc(ws.cell(base_r+2, 2), fill=FILL_PARAM, border=BORDER)
            sc(ws.cell(base_r+2, 3), fill=FILL_PARAM, border=BORDER)
            for ci, yr in enumerate(self.forecast_yrs):
                sc(ws.cell(base_r+2, 4+ci), val=self.scen_z_raw[yr][s],
                   fill=FILL_PARAM, font=FONT_PARAM, align=ALIGN_R,
                   fmt=FMT_NUM6, border=BORDER)
            for gi, grade in enumerate(FROM_BUCKETS):
                r = base_r + 3 + gi
                grade_fill = FILL_LIGHT if grade == "90+" else (FILL_ALT if gi % 2 == 0 else FILL_WHITE)
                ws.row_dimensions[r].height = 15
                sc(ws.cell(r, 1), val=grade,
                   fill=grade_fill,
                   font=Font(name="Arial", bold=True, size=9, color="1F4E79"),
                   align=ALIGN_C, border=BORDER)
                sc(ws.cell(r, 2), val=self.ttc[grade],
                   fill=FILL_PARAM, font=FONT_PARAM,
                   align=ALIGN_R, fmt=FMT_PCT6, border=BORDER)
                sc(ws.cell(r, 3), val=self.rho[grade],
                   fill=FILL_PARAM, font=FONT_PARAM,
                   align=ALIGN_R, fmt=FMT_NUM6, border=BORDER)
                for ci, yr in enumerate(self.forecast_yrs):
                    sc(ws.cell(r, 4+ci), val=self.pd_results[s][grade][yr],
                       fill=row_fill, font=SCEN_Z_FONT[s],
                       align=ALIGN_R, fmt=FMT_PCT2, border=BORDER)
            for sp in range(2):
                ws.row_dimensions[base_r+3+len(FROM_BUCKETS)+sp].height = 5
        note_r = 3 + len(self.scenarios) * BLOCK_ROWS
        for rn, val in enumerate([
            (f"Formula: PD = phi((phi_inv(TTC) - sqrt(rho)*Z) / sqrt(1-rho))  |  "
             f"Shock=+/-{int(self.shock*100)}%  |  LTM={round(self.gdp_ltm,2)}, SD={round(self.gdp_sd,2)}"),
            ("TTC=mean ODR per grade  |  rho=Basel II Retail: 0.03*W+0.16*(1-W)  |  PD format: 0.00%"),
        ], start=0):
            ws.row_dimensions[note_r+rn].height = 13
            sc(ws.cell(note_r+rn, 1), val=val, font=FONT_NOTE, align=ALIGN_L)
            ws.merge_cells(start_row=note_r+rn, start_column=1,
                           end_row=note_r+rn, end_column=4+len(self.forecast_yrs))
        ws.column_dimensions["A"].width = 9
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 22
        for ci in range(len(self.forecast_yrs)):
            ws.column_dimensions[get_column_letter(4+ci)].width = 8

    def _sheet_pd_comparison(self, wb):
        ws = wb.create_sheet("PD Comparison")
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 20
        sc(ws.cell(1, 1), val="Vasicek PD - Scenario Comparison",
           font=FONT_TITLE, align=ALIGN_L)
        ws.merge_cells(start_row=1, start_column=1, end_row=1,
                       end_column=3+len(self.scenarios)*len(self.forecast_yrs))
        ws.row_dimensions[2].height = 6
        ws.row_dimensions[3].height = 16
        for ci, hdr in enumerate(["Grade", "TTC", "rho"]):
            sc(ws.cell(3, 1+ci), val=hdr,
               fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
        for si, s in enumerate(self.scenarios):
            start_col = 4 + si * len(self.forecast_yrs)
            sc(ws.cell(3, start_col), val=s,
               fill=SCEN_HDR_FILL[s],
               font=Font(name="Arial", bold=True, color="FFFFFF", size=9),
               align=ALIGN_C, border=BORDER)
            ws.merge_cells(start_row=3, start_column=start_col,
                           end_row=3, end_column=start_col+len(self.forecast_yrs)-1)
        ws.row_dimensions[4].height = 14
        for ci in range(3):
            sc(ws.cell(4, 1+ci), fill=FILL_DARK, border=BORDER)
        for si, s in enumerate(self.scenarios):
            for ci, yr in enumerate(self.forecast_yrs):
                sc(ws.cell(4, 4+si*len(self.forecast_yrs)+ci), val=str(yr),
                   fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
        for gi, grade in enumerate(FROM_BUCKETS):
            row_e      = 5 + gi
            grade_fill = FILL_LIGHT if grade == "90+" else (FILL_ALT if gi % 2 == 0 else FILL_WHITE)
            ws.row_dimensions[row_e].height = 15
            sc(ws.cell(row_e, 1), val=grade,
               fill=grade_fill, font=FONT_LABEL, align=ALIGN_C, border=BORDER)
            sc(ws.cell(row_e, 2), val=self.ttc[grade],
               fill=FILL_PARAM, font=FONT_PARAM, align=ALIGN_R, fmt=FMT_PCT6, border=BORDER)
            sc(ws.cell(row_e, 3), val=self.rho[grade],
               fill=FILL_PARAM, font=FONT_PARAM, align=ALIGN_R, fmt=FMT_NUM6, border=BORDER)
            for si, s in enumerate(self.scenarios):
                for ci, yr in enumerate(self.forecast_yrs):
                    sc(ws.cell(row_e, 4+si*len(self.forecast_yrs)+ci),
                       val=self.pd_results[s][grade][yr],
                       fill=SCEN_ROW_FILL[s], font=SCEN_Z_FONT[s],
                       align=ALIGN_R, fmt=FMT_PCT2, border=BORDER)
        ws.column_dimensions["A"].width = 9
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 10
        for si in range(len(self.scenarios)):
            for ci in range(len(self.forecast_yrs)):
                ws.column_dimensions[get_column_letter(4+si*len(self.forecast_yrs)+ci)].width = 8

    def _sheet_inputs(self, wb):
        ws = wb.create_sheet("Inputs TTC & rho")
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 20
        sc(ws.cell(1, 1),
           val="Model Inputs - TTC (from ODR) and Asset Correlation (Basel II Retail IRB)",
           font=FONT_TITLE, align=ALIGN_L)
        ws.merge_cells("A1:H1")
        ws.row_dimensions[2].height = 8
        ws.row_dimensions[3].height = 15
        for ci, hdr in enumerate(["Grade","TTC (precise)","TTC (%)",
                                   "rho (exact)","rho (rounded)","W (weight)","0.03*W","0.16*(1-W)"]):
            sc(ws.cell(3, 1+ci), val=hdr,
               fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
        for gi, grade in enumerate(FROM_BUCKETS):
            row_e = 4 + gi
            ttc   = self.ttc[grade]
            rho   = self.rho[grade]
            w     = (1 - np.exp(-35 * ttc)) / (1 - np.exp(-35)) if ttc < 1 else 1.0
            rfill = FILL_ALT if gi % 2 == 0 else FILL_WHITE
            ws.row_dimensions[row_e].height = 15
            for ci, (v, fmt) in enumerate(zip(
                [grade, ttc, ttc, rho, round(rho, 2), w, 0.03*w, 0.16*(1-w)],
                ["@", FMT_NUM6, FMT_PCT2, FMT_NUM6, FMT_NUM2, FMT_NUM6, FMT_NUM6, FMT_NUM6]
            )):
                sc(ws.cell(row_e, 1+ci), val=v, fill=rfill,
                   font=FONT_LABEL if ci == 0 else FONT_BODY,
                   align=ALIGN_C if ci == 0 else ALIGN_R,
                   fmt=fmt, border=BORDER)
        ws.row_dimensions[10].height = 8
        sc(ws.cell(11, 1),
           val=("TTC = mean ODR per grade (active years only)  |  "
                "rho = Basel II Retail IRB: 0.03*W + 0.16*(1-W)  |  "
                "W = (1 - e^(-35*TTC)) / (1 - e^(-35))  |  rho in [0.03, 0.16]"),
           font=FONT_NOTE, align=ALIGN_L)
        ws.merge_cells("A11:H11")
        for ci, w in enumerate([9, 14, 10, 12, 12, 12, 10, 12]):
            ws.column_dimensions[get_column_letter(1+ci)].width = w

    def _sheet_survival(self, wb):
        ws = wb.create_sheet("Survival Analysis")
        ws.sheet_view.showGridLines = False

        sur_bc = {"Base": 1, "Upturn": 12, "Downturn": 23}
        sur_ncols = 10  # label + TTC + 8 years
        n_yrs = len(self.forecast_yrs)

        # Row 1: Title
        ws.row_dimensions[1].height = 22
        sc(ws.cell(1, 1), val="Survival Analysis",
           font=FONT_TITLE, align=ALIGN_L)
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=3 * sur_ncols + 2)
        ws.row_dimensions[2].height = 6

        def _sur_block(top_row, sub_label, data_dict, fmt):
            for scen in self.scenarios:
                bc = sur_bc[scen]
                hdr_fill, hdr_font = FILL_DARK, FONT_HDR

                ws.row_dimensions[top_row].height = 15
                lbl = scen if not sub_label else f"{scen}   {sub_label}"
                sc(ws.cell(top_row, bc), val=lbl,
                   fill=SCEN_HDR_FILL[scen], font=hdr_font, align=ALIGN_C, border=BORDER)
                ws.merge_cells(start_row=top_row, start_column=bc,
                               end_row=top_row, end_column=bc + sur_ncols - 1)

                hdr_row = top_row + 1
                ws.row_dimensions[hdr_row].height = 14
                sc(ws.cell(hdr_row, bc), val="Grades",
                   fill=FILL_MED, font=FONT_HDR, align=ALIGN_C, border=BORDER)
                sc(ws.cell(hdr_row, bc + 1), val="TTC",
                   fill=FILL_MED, font=FONT_HDR, align=ALIGN_C, border=BORDER)
                for di, yr in enumerate(self.forecast_yrs):
                    sc(ws.cell(hdr_row, bc + 2 + di), val=str(yr),
                       fill=FILL_MED, font=FONT_HDR, align=ALIGN_C, border=BORDER)

                for ri, grade in enumerate(FROM_BUCKETS):
                    dr = hdr_row + 1 + ri
                    ws.row_dimensions[dr].height = 14
                    row_fill = FILL_LIGHT if grade == "90+" else (FILL_ALT if ri % 2 == 0 else FILL_WHITE)
                    sc(ws.cell(dr, bc), val=grade,
                       fill=row_fill,
                       font=Font(name="Arial", bold=True, size=9, color="1F4E79"),
                       align=ALIGN_C, border=BORDER)
                    sc(ws.cell(dr, bc + 1), val=self.ttc[grade],
                       fill=FILL_PARAM, font=FONT_PARAM,
                       align=ALIGN_R, fmt=FMT_PCT2, border=BORDER)
                    for di, yr in enumerate(self.forecast_yrs):
                        v = data_dict[scen][grade][yr]
                        sc(ws.cell(dr, bc + 2 + di), val=v,
                           fill=row_fill, font=FONT_BODY,
                           align=ALIGN_R, fmt=fmt, border=BORDER)

            # Spacer columns
            for sp_col in [11, 22]:
                for r in range(top_row, top_row + 7):
                    ws.cell(r, sp_col).fill = FILL_WHITE

        # Draw all 3 sub-tables
        sub_tables = [
            (3, "", self.pd_results, FMT_PCT2),
            (11, "(1-p)", self.surv_1, FMT_PCT2),
            (19, "(1-p)\u2219(1-p)\u2219\u2219\u2219", self.cumul_surv, FMT_PCT2),
        ]

        for top_row, sub_label, data_src, fmt in sub_tables:
            if top_row > 3:
                for sr in range(top_row - 2, top_row):
                    ws.row_dimensions[sr].height = 6
            _sur_block(top_row, sub_label, data_src, fmt)

        # Column widths
        for scen in self.scenarios:
            bc = sur_bc[scen]
            ws.column_dimensions[get_column_letter(bc)].width = 9
            ws.column_dimensions[get_column_letter(bc + 1)].width = 7
            for di in range(n_yrs):
                ws.column_dimensions[get_column_letter(bc + 2 + di)].width = 7
        for sp_col in [11, 22]:
            ws.column_dimensions[get_column_letter(sp_col)].width = 2

        # Note row
        note_r = 26
        ws.row_dimensions[note_r].height = 12
        sc(ws.cell(note_r, 1),
           val=("(1-p) = 1 - Vasicek PD  |  "
                "(1-p)\u00b7(1-p)\u00b7\u00b7\u00b7 = cumulative product of single-period survival probabilities  |  "
                f"Weights - Base:{self.scen_weights['Base']*100:.0f}%  "
                f"Upturn:{self.scen_weights['Upturn']*100:.0f}%  "
                f"Downturn:{self.scen_weights['Downturn']*100:.0f}%"),
           font=FONT_NOTE, align=ALIGN_L)
        ws.merge_cells(start_row=note_r, start_column=1,
                       end_row=note_r, end_column=3 * sur_ncols + 2)

    def _sheet_pit_pd(self, wb):
        ws = wb.create_sheet("PIT PD")
        ws.sheet_view.showGridLines = False

        pit_bc = {"Base": 1, "Upturn": 12, "Downturn": 23}
        pit_ncols = 10

        # Row 1: Title
        ws.row_dimensions[1].height = 22
        sc(ws.cell(1, 1), val="Marginal Probability of Default",
           font=FONT_TITLE, align=ALIGN_L)
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=3 * pit_ncols + 3)
        ws.row_dimensions[2].height = 6

        # Section A: Marginal PD per scenario (rows 3-9)
        for scen in self.scenarios:
            bc = pit_bc[scen]
            ws.row_dimensions[3].height = 15
            sc(ws.cell(3, bc), val=scen,
               fill=SCEN_HDR_FILL[scen],
               font=Font(name="Arial", bold=True, color="FFFFFF", size=9),
               align=ALIGN_C, border=BORDER)
            ws.merge_cells(start_row=3, start_column=bc,
                           end_row=3, end_column=bc + pit_ncols - 1)

            ws.row_dimensions[4].height = 14
            sc(ws.cell(4, bc), val="Grades",
               fill=FILL_MED, font=FONT_HDR, align=ALIGN_C, border=BORDER)
            sc(ws.cell(4, bc + 1), val="TTC",
               fill=FILL_MED, font=FONT_HDR, align=ALIGN_C, border=BORDER)
            for di, yr in enumerate(self.forecast_yrs):
                sc(ws.cell(4, bc + 2 + di), val=str(yr),
                   fill=FILL_MED, font=FONT_HDR, align=ALIGN_C, border=BORDER)

            for ri, grade in enumerate(FROM_BUCKETS):
                dr = 5 + ri
                ws.row_dimensions[dr].height = 14
                row_fill = FILL_LIGHT if grade == "90+" else (FILL_ALT if ri % 2 == 0 else FILL_WHITE)
                sc(ws.cell(dr, bc), val=grade,
                   fill=row_fill,
                   font=Font(name="Arial", bold=True, size=9, color="1F4E79"),
                   align=ALIGN_C, border=BORDER)
                sc(ws.cell(dr, bc + 1), val=self.ttc[grade],
                   fill=FILL_PARAM, font=FONT_PARAM,
                   align=ALIGN_R, fmt=FMT_PCT2, border=BORDER)
                for di, yr in enumerate(self.forecast_yrs):
                    v = self.marginal_pd[scen][grade][yr]
                    sc(ws.cell(dr, bc + 2 + di), val=v,
                       fill=row_fill, font=FONT_BODY,
                       align=ALIGN_R, fmt=FMT_PCT2, border=BORDER)

        # Spacer cols
        for sp_col in [11, 22]:
            ws.column_dimensions[get_column_letter(sp_col)].width = 2

        # Section B: Scenario Weights (rows 10-13)
        ws.row_dimensions[10].height = 8
        ws.row_dimensions[11].height = 16
        sc(ws.cell(11, 1), val="Scenario Weights",
           fill=FILL_DARK, font=FONT_TITLE, align=ALIGN_L)
        ws.merge_cells(start_row=11, start_column=1, end_row=11, end_column=6)

        ws.row_dimensions[12].height = 14
        ws.row_dimensions[13].height = 14
        for ci, scen in enumerate(self.scenarios):
            sc(ws.cell(12, 1 + ci), val=scen,
               fill=SCEN_HDR_FILL[scen],
               font=Font(name="Arial", bold=True, color="FFFFFF", size=9),
               align=ALIGN_C, border=BORDER)
            sc(ws.cell(13, 1 + ci), val=self.scen_weights[scen],
               fill=SCEN_ROW_FILL[scen], font=FONT_PARAM,
               align=ALIGN_C, fmt="0.00%", border=BORDER)

        # Section C: PIT PD (rows 15-21)
        ws.row_dimensions[14].height = 8
        ws.row_dimensions[15].height = 18
        sc(ws.cell(15, 1), val="Point-In-Time Probability of Default  (PIT PD)",
           font=FONT_TITLE, align=ALIGN_L)
        ws.merge_cells(start_row=15, start_column=1,
                       end_row=15, end_column=3 * pit_ncols + 3)

        ws.row_dimensions[16].height = 14
        for ci, hdr in enumerate(["Grades", "TTC"] + [str(y) for y in self.forecast_yrs]):
            sc(ws.cell(16, 1 + ci), val=hdr,
               fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)

        for ri, grade in enumerate(FROM_BUCKETS):
            dr = 17 + ri
            ws.row_dimensions[dr].height = 14
            row_fill = FILL_LIGHT if grade == "90+" else (FILL_ALT if ri % 2 == 0 else FILL_WHITE)
            sc(ws.cell(dr, 1), val=grade,
               fill=row_fill,
               font=Font(name="Arial", bold=True, size=9, color="1F4E79"),
               align=ALIGN_C, border=BORDER)
            sc(ws.cell(dr, 2), val=self.ttc[grade],
               fill=FILL_PARAM, font=FONT_PARAM,
               align=ALIGN_R, fmt=FMT_PCT2, border=BORDER)
            for di, yr in enumerate(self.forecast_yrs):
                v = self.pit_pd_vals[grade][yr]
                sc(ws.cell(dr, 3 + di), val=v,
                   fill=row_fill, font=FONT_BODY,
                   align=ALIGN_R, fmt=FMT_PCT2, border=BORDER)

        # Note row
        note_r = 23
        ws.row_dimensions[22].height = 8
        ws.row_dimensions[note_r].height = 12
        sc(ws.cell(note_r, 1),
           val=("PIT PD = \u03a3(scenario weight \u00d7 marginal PD)  |  "
                "Marginal PD t\u2081 = Vasicek PD t\u2081;  "
                "Marginal PD t\u2099 = S(t\u2099\u208b\u2081) \u2212 S(t\u2099)"),
           font=FONT_NOTE, align=ALIGN_L)
        ws.merge_cells(start_row=note_r, start_column=1,
                       end_row=note_r, end_column=3 * pit_ncols + 3)

        # Column widths
        for scen in self.scenarios:
            bc = pit_bc[scen]
            ws.column_dimensions[get_column_letter(bc)].width = 9
            ws.column_dimensions[get_column_letter(bc + 1)].width = 7
            for di in range(len(self.forecast_yrs)):
                ws.column_dimensions[get_column_letter(bc + 2 + di)].width = 7
        ws.column_dimensions["A"].width = 9
        ws.column_dimensions["B"].width = 7
        for di in range(len(self.forecast_yrs)):
            ws.column_dimensions[get_column_letter(3 + di)].width = 7

    # ── Collect results for API ──────────────────────────────────────────

    def _collect_results(self):
        odr_summary = []
        for yr in self.year_pairs:
            res = self.odr_results[yr]
            nm  = res["months"]
            odr_summary.append({
                "period":    f"{yr}-{yr+1}",
                "from_yr":   yr,
                "odr":       None if res["odr"] is None else round(res["odr"], 6),
                "months":    nm,
                "total_obs": res["total_obs"],
                "status":    "no_data" if nm == 0 else ("partial" if nm < 12 else "full"),
            })

        ttc_rho = [
            {"grade": fb, "ttc": round(self.ttc[fb], 6), "rho": round(self.rho[fb], 6)}
            for fb in FROM_BUCKETS
        ]

        mav_list = [
            {"mev": SERIES_MAP[c], "code": c,
             "ltm": self.mav_params[c]["LTM"], "sd": self.mav_params[c]["SD"],
             "cv": round(abs(self.mav_params[c]["SD"] / self.mav_params[c]["LTM"]), 4)
                   if self.mav_params[c]["LTM"] != 0 else None}
            for c in SERIES_ORDER
        ]

        z_dict = {}
        for code in SERIES_ORDER:
            z_dict[SERIES_MAP[code]] = {
                str(yr): self.z_factors[code][yr] for yr in self.display_years
            }

        scen_data = {"years": [str(yr) for yr in self.display_years]}
        for s in self.scenarios:
            scen_data[s] = [self.mev_scenarios[yr][s] for yr in self.display_years]

        vasicek = {"years": [str(yr) for yr in self.forecast_yrs]}
        for s in self.scenarios:
            vasicek[s] = {}
            for grade in FROM_BUCKETS:
                vasicek[s][grade] = [
                    round(self.pd_results[s][grade][yr], 6) for yr in self.forecast_yrs
                ]

        # ODR matrices for transition matrix viewer
        odr_matrices = {}
        for yr in self.year_pairs:
            res = self.odr_results[yr]
            if res["months"] == 0:
                continue
            mat = res["matrix"]
            period = f"{yr}-{yr+1}"
            odr_matrices[period] = {
                fb: {tb: mat[fb][tb] for tb in TO_BUCKETS}
                for fb in FROM_BUCKETS
            }

        # ODR by grade for heatmap
        odr_by_grade = {}
        for fb in FROM_BUCKETS:
            odr_by_grade[fb] = []
            for yr in self.year_pairs:
                res = self.odr_results[yr]
                if res["months"] == 0:
                    continue
                mat = res["matrix"]
                row = mat[fb]
                total = sum(row.values())
                dft = sum(row[d] for d in DEFAULT_TO_B) if total else 0
                odr_by_grade[fb].append({
                    "period": f"{yr}-{yr+1}",
                    "odr": round(dft / total, 6) if total else 0,
                })

        # Raw GDP Z for sensitivity analysis
        gdp_z = {str(yr): round(self.gdp_z_raw[yr], 6) for yr in self.forecast_yrs}

        # Correlation curve for chart
        corr_curve = []
        for pd_bps in list(range(1, 100, 2)) + list(range(100, 10001, 50)):
            pd_val = pd_bps / 10000
            corr_curve.append({
                "pd": round(pd_val, 4),
                "rho": round(basel_retail_rho(pd_val), 6),
            })

        return {
            "odr_summary":   odr_summary,
            "ttc_rho":       ttc_rho,
            "mav_params":    mav_list,
            "z_factors":     z_dict,
            "scenarios":     scen_data,
            "vasicek_pd":    vasicek,
            "odr_matrices":  odr_matrices,
            "odr_by_grade":  odr_by_grade,
            "gdp_z_raw":     gdp_z,
            "corr_curve":    corr_curve,
            # Survival analysis
            "survival_1": {
                s: {
                    grade: [round(self.surv_1[s][grade][yr], 6) for yr in self.forecast_yrs]
                    for grade in FROM_BUCKETS
                }
                for s in self.scenarios
            },
            "cumul_survival": {
                s: {
                    grade: [round(self.cumul_surv[s][grade][yr], 6) for yr in self.forecast_yrs]
                    for grade in FROM_BUCKETS
                }
                for s in self.scenarios
            },
            # Marginal PD
            "marginal_pd": {
                s: {
                    grade: [round(self.marginal_pd[s][grade][yr], 6) for yr in self.forecast_yrs]
                    for grade in FROM_BUCKETS
                }
                for s in self.scenarios
            },
            # PIT PD
            "pit_pd": {
                grade: [round(self.pit_pd_vals[grade][yr], 6) for yr in self.forecast_yrs]
                for grade in FROM_BUCKETS
            },
            "lifetime_pd": {
                grade: round(self.lifetime_pd[grade], 6) for grade in FROM_BUCKETS
            },
            "scen_weights": self.scen_weights,
            "config_used": {
                "shock":         self.shock,
                "tm_start_year": self.tm_start_year,
                "hist_cutoff":   self.hist_cutoff,
                "gdp_ltm":       round(self.gdp_ltm, 2),
                "gdp_sd":        round(self.gdp_sd, 2),
            },
        }

    # ── Public entry point ───────────────────────────────────────────────

    def run(self):
        self._load_dpd()
        self._compute_odr()
        self._compute_ttc_rho()
        self._load_weo()
        self._compute_mav()
        self._compute_scenarios()
        self._compute_vasicek()
        self._compute_survival()
        self._compute_pit_pd()
        self._generate_excel()
        return self._collect_results()
