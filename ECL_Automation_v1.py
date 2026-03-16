#--CreditRisk ECL Automation v1
#----------------------------------------------------------------------------------

import re
import numpy as np
import pandas as pd
from collections import defaultdict
from scipy.stats import norm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# ── INPUTS 
# ══════════════════════════════════════════════════════════════════════════════

DPD_FILE    = "dummy_dpd_excel.xlsx"
WEO_FILE    = "WEO_Data.xlsx"
OUTPUT_FILE = "V7_ECL_Output.xlsx"

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION  (model parameters)
# ══════════════════════════════════════════════════════════════════════════════

# — Transition Matrix —
TM_START_YEAR = 2020           

# — DPD / ODR —
FROM_STATES  = ["0", "1", "2", "3", "4"]
TO_STATES    = ["0", "1", "2", "3", "4", "WO", "ARC", "Closed"]
FROM_BUCKETS = ["0", "1-30", "31-60", "61-90", "90+"]
TO_BUCKETS   = ["0", "1-30", "31-60", "61-90", "90+", "WO", "ARC", "Closed"]
DEFAULT_TO   = {"4", "WO", "ARC"}         
DEFAULT_TO_B = {"90+", "WO", "ARC"}       

MONTH_MAP = {
    "jan": 1, "feb": 2,  "mar": 3,  "apr": 4,
    "may": 5, "jun": 6,  "jul": 7,  "aug": 8,
    "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    "june": 6, "july": 7,
}
MONTH_LABELS = {
    1:"Jan", 2:"Feb",  3:"Mar",  4:"Apr",
    5:"May", 6:"Jun",  7:"Jul",  8:"Aug",
    9:"Sep", 10:"Oct", 11:"Nov", 12:"Dec",
}
DPD_BUCKET = {
    "0":"0", "1":"1-30", "2":"31-60", "3":"61-90", "4":"90+",
    "WO":"WO", "ARC":"ARC", "Closed":"Closed",
}

# — WEO / MAV —
SERIES_MAP = {
    "LUR":         "Unemployment Rate",
    "NGDP_RPCH":   "GDP",
    "PCPIPCH":     "Inflation",
    "GGX_NGDP":    "Govt_Expenditure",
    "GGXWDG_NGDP": "General government gross debt",
}
SERIES_ORDER  = ["LUR", "NGDP_RPCH", "PCPIPCH", "GGX_NGDP", "GGXWDG_NGDP"]
GDP_CODE      = "NGDP_RPCH"
WEO_YEARS     = list(range(2019, 2028))
EXTRAP_YEARS  = list(range(2028, 2033))
TREND_WINDOW  = [2025, 2026, 2027]
DISPLAY_YEARS = WEO_YEARS + EXTRAP_YEARS   # 2019-2032
HIST_CUTOFF   = 2024
CALIB_YEARS   = list(range(2019, 2026))    # LTM/SD window n=7
GDP_PRECISE   = {2025: 6.198, 2026: 6.265, 2027: 6.471}

# — Vasicek / Scenarios —
SCENARIOS    = ["Base", "Upturn", "Downturn"]
SHOCK        = 0.10                        # ±10% of |Z|
FORECAST_YRS = list(range(2025, 2033))

# ══════════════════════════════════════════════════════════════════════════════
# STYLES  (defined once, shared by all sheets)
# ══════════════════════════════════════════════════════════════════════════════

FILL_DARK    = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
FILL_MED     = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
FILL_LIGHT   = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
FILL_ALT     = PatternFill("solid", start_color="EBF5FB", end_color="EBF5FB")
FILL_WHITE   = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
FILL_GREEN   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
FILL_WARN    = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
FILL_RED     = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
FILL_GREY    = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
FILL_GOLD    = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
FILL_HIST    = PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1")
FILL_FCST    = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
FILL_EXTRAP  = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
FILL_BASE    = PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1")
FILL_UP      = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
FILL_DOWN    = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
FILL_PARAM   = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")

SCEN_HDR_FILL = {
    "Base":     PatternFill("solid", start_color="2E75B6", end_color="2E75B6"),
    "Upturn":   PatternFill("solid", start_color="375623", end_color="375623"),
    "Downturn": PatternFill("solid", start_color="C00000", end_color="C00000"),
}
SCEN_ROW_FILL = {"Base": FILL_BASE, "Upturn": FILL_UP, "Downturn": FILL_DOWN}

FONT_TITLE   = Font(name="Arial", bold=True,  color="1F4E79", size=12)
FONT_HDR     = Font(name="Arial", bold=True,  color="FFFFFF", size=9)
FONT_LABEL   = Font(name="Arial", bold=True,  color="1F4E79", size=9)
FONT_BODY    = Font(name="Arial", size=9)
FONT_TOTAL   = Font(name="Arial", bold=True,  size=9)
FONT_PARAM   = Font(name="Arial", bold=True,  color="7F4F00", size=9)
FONT_SECTION = Font(name="Arial", bold=True,  color="FFFFFF", size=10)
FONT_ODR     = Font(name="Arial", bold=True,  color="1F4E79", size=9)
FONT_GREY    = Font(name="Arial", italic=True, color="808080", size=9)
FONT_NA      = Font(name="Arial", bold=True,  color="C00000", size=9)
FONT_NOTE    = Font(name="Arial", italic=True, color="595959", size=8)
FONT_Z_POS   = Font(name="Arial", size=9, color="375623")
FONT_Z_NEG   = Font(name="Arial", size=9, color="C00000")
FONT_BASE    = Font(name="Arial", size=9, color="1F4E79")
FONT_UP      = Font(name="Arial", size=9, color="375623")
FONT_DOWN    = Font(name="Arial", size=9, color="C00000")
SCEN_Z_FONT  = {"Base": FONT_BASE, "Upturn": FONT_UP, "Downturn": FONT_DOWN}

_thin  = Side(style="thin", color="B0C4DE")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

ALIGN_C = Alignment(horizontal="center", vertical="center")
ALIGN_R = Alignment(horizontal="right",  vertical="center")
ALIGN_L = Alignment(horizontal="left",   vertical="center")

FMT_INT  = "#,##0"
FMT_PCT2 = "0.00%"
FMT_PCT6 = "0.000000%"
FMT_NUM2 = "0.00"
FMT_NUM6 = "0.000000"
FMT_DATE = "YYYY-MM-DD"


# ══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS  (defined once, used everywhere)
# ══════════════════════════════════════════════════════════════════════════════

def sc(cell, fill=None, font=None, align=None, fmt=None, border=None, val=None):
    """Apply all styles and optional value to an openpyxl cell in one call."""
    if val is not None:  cell.value         = val
    if fill:             cell.fill          = fill
    if font:             cell.font          = font
    if align:            cell.alignment     = align
    if fmt:              cell.number_format = fmt
    if border:           cell.border        = border


def parse_dpd_col(col: str):
    """Return (month_int, year_int) from 'DPD_Apr-20', else None."""
    m = re.match(r"DPD_([A-Za-z]+)-(\d{2})$", col)
    if not m:
        return None
    month = MONTH_MAP.get(m.group(1).lower())
    return (month, 2000 + int(m.group(2))) if month else None


def cast_dpd(v):
    """Normalise a DPD cell value to string integer or string label."""
    if pd.isna(v):
        return np.nan
    try:
        return str(int(float(v)))
    except (ValueError, TypeError):
        return str(v)


def linear_extrap(series: dict, horizon: int) -> float:
    """OLS linear extrapolation from TREND_WINDOW to a future year."""
    base = TREND_WINDOW[0]
    x    = np.array([yr - base for yr in TREND_WINDOW], dtype=float)
    y    = np.array([series[yr] for yr in TREND_WINDOW], dtype=float)
    slope, intercept = np.polyfit(x, y, 1)
    return round(intercept + slope * (horizon - base), 4)


def year_fill(yr: int) -> PatternFill:
    if yr <= HIST_CUTOFF: return FILL_HIST
    if yr <= 2027:        return FILL_FCST
    return FILL_EXTRAP


def year_font(yr: int) -> Font:
    if yr <= HIST_CUTOFF: return Font(name="Arial", size=9, color="1F4E79")
    if yr <= 2027:        return Font(name="Arial", size=9, color="375623")
    return Font(name="Arial", size=9, color="7F4F00", italic=True)


def basel_retail_rho(pd_val: float) -> float:
    """Basel II Retail IRB asset correlation (BCBS §328): ρ ∈ [0.03, 0.16]."""
    if pd_val >= 1.0:
        return 0.03
    w = (1 - np.exp(-35 * pd_val)) / (1 - np.exp(-35))
    return 0.03 * w + 0.16 * (1 - w)


def vasicek_pd(ttc: float, rho: float, z: float) -> float:
    """Vasicek single-factor PD: Φ((Φ⁻¹(TTC) − √ρ·Z) / √(1−ρ))."""
    if ttc >= 1.0:
        return 1.0
    return norm.cdf((norm.ppf(ttc) - np.sqrt(rho) * z) / np.sqrt(1 - rho))


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — LOAD DPD DATA
# ══════════════════════════════════════════════════════════════════════════════

print("Loading DPD data...")
df_dpd = pd.read_excel(DPD_FILE)
df_dpd.columns = df_dpd.columns.str.strip()

col_lookup: dict[tuple, str] = {
    parse_dpd_col(c): c for c in df_dpd.columns if parse_dpd_col(c)
}

# TM pairs: for each month, list of (from_yr, to_yr) starting at TM_START_YEAR
tm_pairs: dict[int, list] = {}
for month in range(1, 13):
    pairs, yr = [], TM_START_YEAR
    while (month, yr) in col_lookup and (month, yr + 1) in col_lookup:
        pairs.append((yr, yr + 1))
        yr += 1
    tm_pairs[month] = pairs

# ODR pairs: all years where at least one month has active loans
pairs_by_yr: dict[int, list] = defaultdict(list)
for (m, yr) in col_lookup:
    if (m, yr + 1) in col_lookup:
        pairs_by_yr[yr].append(m)
YEAR_PAIRS = sorted(pairs_by_yr.keys())


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — TRANSITION MATRIX COMPUTATION
# ══════════════════════════════════════════════════════════════════════════════

def compute_transition(from_col: str, to_col: str) -> dict:
    """Return from-state keyed dict of {to_state: count, Total: int, Default: float}."""
    frm = df_dpd[from_col].map(cast_dpd)
    to_ = df_dpd[to_col].map(cast_dpd)
    rows = {}
    for fs in FROM_STATES:
        mask   = frm == fs
        subset = to_[mask]
        total  = int(mask.sum())
        counts = {ts: int((subset == ts).sum()) for ts in TO_STATES}
        dft    = sum(counts[d] for d in DEFAULT_TO)
        rows[fs] = {**counts, "Total": total,
                    "Default": dft / total if total else 0.0}
    # Total row
    agg   = {ts: sum(rows[fs][ts] for fs in FROM_STATES) for ts in TO_STATES}
    grand = sum(rows[fs]["Total"] for fs in FROM_STATES)
    agg["Total"]   = grand
    agg["Default"] = sum(agg[d] for d in DEFAULT_TO) / grand if grand else 0.0
    rows["Total"]  = agg
    return rows


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — ODR COMPUTATION
# ══════════════════════════════════════════════════════════════════════════════

def compute_odr_matrix(from_yr: int):
    """Aggregate all monthly transitions for one annual period."""
    agg         = {fb: {tb: 0 for tb in TO_BUCKETS} for fb in FROM_BUCKETS}
    months_used = 0
    total_obs   = 0

    for m in sorted(pairs_by_yr[from_yr]):
        frm_s = df_dpd[col_lookup[(m, from_yr)]].map(cast_dpd).map(
                    lambda v: DPD_BUCKET.get(v, np.nan))
        to_s  = df_dpd[col_lookup[(m, from_yr + 1)]].map(cast_dpd).map(
                    lambda v: DPD_BUCKET.get(v, np.nan))
        obs = int(frm_s.isin(FROM_BUCKETS).sum())
        if obs == 0:
            continue
        months_used += 1
        total_obs   += obs
        for fb in FROM_BUCKETS:
            mask = frm_s == fb
            sub  = to_s[mask]
            for tb in TO_BUCKETS:
                agg[fb][tb] += int((sub == tb).sum())

    return agg, months_used, total_obs


odr_results = {}
for from_yr in YEAR_PAIRS:
    mat, nm, total_obs = compute_odr_matrix(from_yr)
    tot   = {tb: sum(mat[fb][tb] for fb in FROM_BUCKETS) for tb in TO_BUCKETS}
    grand = sum(tot.values())
    odr   = sum(tot[d] for d in DEFAULT_TO_B) / grand if grand else None
    odr_results[from_yr] = {
        "matrix": mat, "odr": odr,
        "months": nm,  "total_obs": total_obs,
    }

print("ODR summary:")
for yr, res in odr_results.items():
    odr_str = f"{res['odr']*100:.4f}%" if res["odr"] is not None else "N/A"
    nm = res["months"]
    note = ("← No active loans" if nm == 0 else
            f"← Partial ({nm} months)" if nm < 12 else
            f"← Few obs ({res['total_obs']})" if res["total_obs"] < 100 else "")
    print(f"  {yr}→{yr+1}: {odr_str:>9}  {note}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — TTC AND ASSET CORRELATION
# ══════════════════════════════════════════════════════════════════════════════

odr_by_grade: dict[str, list] = {fb: [] for fb in FROM_BUCKETS}
for yr in YEAR_PAIRS:
    mat, nm, _ = compute_odr_matrix(yr)
    if nm == 0:
        continue
    for fb in FROM_BUCKETS:
        row   = mat[fb]
        total = sum(row.values())
        if total == 0:
            continue
        dft = sum(row[d] for d in DEFAULT_TO_B)
        odr_by_grade[fb].append(dft / total)

TTC: dict[str, float] = {
    fb: float(np.mean(odr_by_grade[fb])) if odr_by_grade[fb] else 1.0
    for fb in FROM_BUCKETS
}
TTC["90+"] = 1.0
RHO: dict[str, float] = {fb: basel_retail_rho(TTC[fb]) for fb in FROM_BUCKETS}

print("\nTTC & ρ per grade:")
for fb in FROM_BUCKETS:
    print(f"  {fb:<8}  TTC={TTC[fb]*100:.4f}%  ρ={RHO[fb]:.6f}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 5 — WEO DATA → MAV VALUES + LTM/SD
# ══════════════════════════════════════════════════════════════════════════════

print("\nLoading WEO data...")
df_weo = pd.read_excel(WEO_FILE, header=None)

weo_year_cols: dict[int, int] = {}
for c, v in enumerate(df_weo.iloc[0]):
    try:
        yr = int(float(v))
        if 2000 <= yr <= 2027:
            weo_year_cols[yr] = c
    except (ValueError, TypeError):
        pass

raw_weo: dict[str, dict[int, float]] = {}
for _, row in df_weo.iterrows():
    code = str(row[1]).strip()
    if code in SERIES_MAP:
        raw_weo[code] = {yr: float(row[weo_year_cols[yr]]) for yr in WEO_YEARS}

MAV: dict[str, dict[int, float]] = {}
for code in SERIES_ORDER:
    MAV[code] = {}
    for yr in WEO_YEARS:
        MAV[code][yr] = round(raw_weo[code][yr], 4)
    for yr in EXTRAP_YEARS:
        MAV[code][yr] = linear_extrap(raw_weo[code], yr)

mav_params: dict[str, dict] = {}
for code in SERIES_ORDER:
    vals = np.array([MAV[code][yr] for yr in CALIB_YEARS])
    mav_params[code] = {
        "LTM": round(float(vals.mean()),      2),
        "SD":  round(float(vals.std(ddof=0)), 2),
    }

z_factors: dict[str, dict[int, float]] = {}
for code in SERIES_ORDER:
    ltm = mav_params[code]["LTM"]
    sd  = mav_params[code]["SD"]
    z_factors[code] = {
        yr: round((GDP_PRECISE.get(yr, MAV[code][yr]) if code == GDP_CODE
                   else MAV[code][yr]) / sd - ltm / sd, 2)
        for yr in DISPLAY_YEARS
    }

print(f"MAV calibration (window {CALIB_YEARS[0]}–{CALIB_YEARS[-1]}, population std):")
for code in SERIES_ORDER:
    p = mav_params[code]
    print(f"  {SERIES_MAP[code]:<35}  LTM={p['LTM']:.2f}  SD={p['SD']:.2f}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 6 — GDP Z-FACTOR SCENARIOS
# ══════════════════════════════════════════════════════════════════════════════

gdp_ltm = mav_params[GDP_CODE]["LTM"]
gdp_sd  = mav_params[GDP_CODE]["SD"]

gdp_z_raw: dict[int, float] = {
    yr: (GDP_PRECISE.get(yr, MAV[GDP_CODE][yr]) - gdp_ltm) / gdp_sd
    for yr in DISPLAY_YEARS
}

mev_scenarios: dict[int, dict[str, float]] = {
    yr: {
        "Base":     round(z,                  2),
        "Upturn":   round(z + abs(z) * SHOCK, 2),
        "Downturn": round(z - abs(z) * SHOCK, 2),
    }
    for yr, z in gdp_z_raw.items()
}


# ══════════════════════════════════════════════════════════════════════════════
# STEP 7 — VASICEK PD
# ══════════════════════════════════════════════════════════════════════════════

# Use unrounded Z for Vasicek (more precise)
scen_z_raw: dict[int, dict[str, float]] = {
    yr: {
        "Base":     gdp_z_raw[yr],
        "Upturn":   gdp_z_raw[yr] + abs(gdp_z_raw[yr]) * SHOCK,
        "Downturn": gdp_z_raw[yr] - abs(gdp_z_raw[yr]) * SHOCK,
    }
    for yr in FORECAST_YRS
}

pd_results: dict[str, dict[str, dict[int, float]]] = {}
for scen in SCENARIOS:
    pd_results[scen] = {
        grade: {
            yr: vasicek_pd(TTC[grade], RHO[grade], scen_z_raw[yr][scen])
            for yr in FORECAST_YRS
        }
        for grade in FROM_BUCKETS
    }

print("\nVasicek PD — Base:")
for grade in FROM_BUCKETS:
    pds = "  ".join(f"{pd_results['Base'][grade][yr]*100:.2f}%" for yr in FORECAST_YRS)
    print(f"  {grade:<8}  {pds}")


# ══════════════════════════════════════════════════════════════════════════════
# BUILD SINGLE EXCEL WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════

wb = Workbook()

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1 — TRANSITION MATRIX
# ─────────────────────────────────────────────────────────────────────────────

ws_tm = wb.active
ws_tm.title = "Transition Matrix"
ws_tm.sheet_view.showGridLines = False

TM_COLS_PER_YEAR = 12
TM_ROWS_PER_TM   = 10
TO_KEYS    = ["0","1","2","3","4","WO","ARC","Closed","Total","Default"]
COL_LABELS = ["0","1","2.0","3.0","4.0","WO","ARC","Closed","Total","Defualt"]

max_years = max((len(p) for p in tm_pairs.values()), default=0)
for yi in range(max_years):
    bc = yi * TM_COLS_PER_YEAR + 1
    ws_tm.column_dimensions[get_column_letter(bc)].width      = 2
    ws_tm.column_dimensions[get_column_letter(bc + 1)].width  = 8
    ws_tm.column_dimensions[get_column_letter(bc + 2)].width  = 12
    ws_tm.column_dimensions[get_column_letter(bc + 3)].width  = 8
    for dc in range(4, 10):
        ws_tm.column_dimensions[get_column_letter(bc + dc)].width = 9
    ws_tm.column_dimensions[get_column_letter(bc + 10)].width = 10
    ws_tm.column_dimensions[get_column_letter(bc + 11)].width = 11

for month in range(1, 13):
    pairs = tm_pairs[month]
    if not pairs:
        continue
    base_row = (month - 1) * TM_ROWS_PER_TM + 1
    ws_tm.row_dimensions[base_row].height     = 15
    ws_tm.row_dimensions[base_row + 1].height = 15
    for ri in range(6):
        ws_tm.row_dimensions[base_row + 2 + ri].height = 14

    for yi, (from_yr, to_yr) in enumerate(pairs):
        bc  = yi * TM_COLS_PER_YEAR + 1
        c1  = bc + 1
        from_date = pd.Timestamp(f"{from_yr}-{month:02d}-01")
        to_date   = pd.Timestamp(f"{to_yr}-{month:02d}-01")

        sc(ws_tm.cell(base_row, c1),     fill=FILL_DARK, font=FONT_HDR,
           align=ALIGN_C, border=BORDER, val=f"TM{month}")
        sc(ws_tm.cell(base_row, c1 + 1), fill=FILL_DARK, font=FONT_HDR,
           align=ALIGN_C, fmt=FMT_DATE,  border=BORDER, val=to_date)
        sc(ws_tm.cell(base_row, c1 + 2), fill=FILL_DARK, font=FONT_HDR,
           align=ALIGN_C, border=BORDER, val=f"Year {yi + 1}")
        for dc in range(3, 11):
            sc(ws_tm.cell(base_row, c1 + dc), fill=FILL_DARK, border=BORDER)

        sc(ws_tm.cell(base_row + 1, c1), fill=FILL_MED, font=FONT_HDR,
           align=ALIGN_C, fmt=FMT_DATE,  border=BORDER, val=from_date)
        for di, lbl in enumerate(COL_LABELS):
            sc(ws_tm.cell(base_row + 1, c1 + 1 + di), fill=FILL_MED, font=FONT_HDR,
               align=ALIGN_C, border=BORDER, val=lbl)

        mat = compute_transition(col_lookup[(month, from_yr)],
                                  col_lookup[(month, to_yr)])
        for ri, fs in enumerate(FROM_STATES + ["Total"]):
            r        = base_row + 2 + ri
            is_total = fs == "Total"
            row_fill = FILL_LIGHT if is_total else (FILL_ALT if ri % 2 == 0 else FILL_WHITE)
            row_font = FONT_TOTAL if is_total else FONT_BODY
            sc(ws_tm.cell(r, c1), fill=row_fill, font=row_font,
               align=ALIGN_C, border=BORDER, val="Total" if is_total else fs)
            for di, key in enumerate(TO_KEYS):
                val = mat[fs][key]
                fmt = FMT_PCT2 if key == "Default" else FMT_INT
                sc(ws_tm.cell(r, c1 + 1 + di), fill=row_fill, font=row_font,
                   align=ALIGN_R, fmt=fmt, border=BORDER, val=val)

ws_tm.freeze_panes = "B3"
print("\nSheet 1: Transition Matrix ✓")


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 — ODR
# ─────────────────────────────────────────────────────────────────────────────

ws_odr = wb.create_sheet("ODR")
ws_odr.sheet_view.showGridLines = False

NUM_ODR_YEARS = len(YEAR_PAIRS)
ODR_COLS_PER_YEAR = 12

ws_odr.row_dimensions[1].height = 22
sc(ws_odr.cell(1, 1), val="Observed Default Rate (ODR) — Full History",
   font=FONT_TITLE, align=ALIGN_L)
ws_odr.row_dimensions[2].height = 6

for yi, from_yr in enumerate(YEAR_PAIRS):
    bc      = yi * ODR_COLS_PER_YEAR + 1
    res     = odr_results[from_yr]
    mat     = res["matrix"]
    nm      = res["months"]
    yn      = yi + 1
    no_data = nm == 0
    partial = 0 < nm < 12

    if no_data:
        year_label = f"Year {yn}  ({from_yr}→{from_yr+1})  [No active loans]"
        hdr_fill   = FILL_GREY
        hdr_font   = Font(name="Arial", bold=True, color="808080", size=9)
    elif partial:
        year_label = f"Year {yn}  ({from_yr}→{from_yr+1})  [{nm} months] *"
        hdr_fill   = FILL_WARN
        hdr_font   = Font(name="Arial", bold=True, color="7F4F00", size=9)
    else:
        year_label = f"Year {yn}  ({from_yr}→{from_yr+1})"
        hdr_fill, hdr_font = FILL_DARK, FONT_HDR

    ws_odr.row_dimensions[3].height = 15
    sc(ws_odr.cell(3, bc + 1), val=year_label,
       fill=hdr_fill, font=hdr_font, align=ALIGN_C, border=BORDER)
    ws_odr.merge_cells(start_row=3, start_column=bc+1, end_row=3, end_column=bc+11)

    ws_odr.row_dimensions[4].height = 15
    sc(ws_odr.cell(4, bc + 1), fill=FILL_MED, border=BORDER)
    for di, tb in enumerate(TO_BUCKETS + ["Total", "Default"]):
        sc(ws_odr.cell(4, bc + 2 + di), val=tb,
           fill=FILL_MED, font=FONT_HDR, align=ALIGN_C, border=BORDER)

    for ri, fb in enumerate(FROM_BUCKETS + ["Total"]):
        row_e    = 5 + ri
        is_total = fb == "Total"
        ws_odr.row_dimensions[row_e].height = 14

        if no_data:
            row_fill, row_font = FILL_GREY, FONT_GREY
        else:
            row_fill = FILL_LIGHT if is_total else (FILL_ALT if ri % 2 == 0 else FILL_WHITE)
            row_font = FONT_TOTAL if is_total else FONT_BODY

        sc(ws_odr.cell(row_e, bc + 1), val=fb, fill=row_fill,
           font=Font(name="Arial", bold=True, size=9,
                     color="808080" if no_data else "1F4E79"),
           align=ALIGN_C, border=BORDER)

        if no_data:
            for di in range(10):
                sc(ws_odr.cell(row_e, bc + 2 + di), val="N/A",
                   fill=FILL_GREY, font=FONT_GREY, align=ALIGN_C, border=BORDER)
            continue

        row_data  = ({tb: sum(mat[b][tb] for b in FROM_BUCKETS) for tb in TO_BUCKETS}
                     if is_total else mat[fb])
        total_val = sum(row_data.values())
        dft_val   = sum(row_data[d] for d in DEFAULT_TO_B) / total_val if total_val else 0.0

        for di, tk in enumerate(TO_BUCKETS + ["Total", "Default"]):
            col_e = bc + 2 + di
            if tk == "Total":
                v, fmt = total_val, FMT_INT
            elif tk == "Default":
                v, fmt = dft_val, FMT_PCT2
            else:
                v, fmt = row_data[tk], FMT_INT
            sc(ws_odr.cell(row_e, col_e), val=v,
               fill=row_fill, font=row_font, align=ALIGN_R, fmt=fmt, border=BORDER)

ws_odr.row_dimensions[11].height = 8
ws_odr.row_dimensions[12].height = 8

ws_odr.row_dimensions[13].height = 15
sc(ws_odr.cell(13, 1), val="Period",
   fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
for yi, from_yr in enumerate(YEAR_PAIRS):
    sc(ws_odr.cell(13, 2 + yi), val=f"{from_yr}→{from_yr+1}",
       fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)

ws_odr.row_dimensions[14].height = 15
sc(ws_odr.cell(14, 1), val="ODR",
   fill=FILL_GREEN, font=FONT_ODR, align=ALIGN_C, border=BORDER)
for yi, from_yr in enumerate(YEAR_PAIRS):
    res     = odr_results[from_yr]
    no_data = res["months"] == 0
    partial = 0 < res["months"] < 12
    fill    = FILL_RED if no_data else (FILL_WARN if partial else FILL_GREEN)
    if no_data:
        sc(ws_odr.cell(14, 2 + yi), val="N/A",
           fill=fill, font=FONT_NA, align=ALIGN_C, border=BORDER)
    else:
        sc(ws_odr.cell(14, 2 + yi), val=res["odr"],
           fill=fill, font=FONT_ODR, align=ALIGN_R, fmt=FMT_PCT2, border=BORDER)

for row_n, (text, colour, italic) in enumerate([
    ("Notes:", "1F4E79", False),
    ("  Red (N/A) — No active loans; ODR cannot be computed.", "C00000", True),
    ("  Amber (*) — Partial year; ODR may be understated.",    "7F4F00", True),
], start=15):
    ws_odr.row_dimensions[row_n].height = 12
    sc(ws_odr.cell(row_n, 1), val=text,
       font=Font(name="Arial", size=8, italic=italic, color=colour), align=ALIGN_L)
    ws_odr.merge_cells(start_row=row_n, start_column=1,
                       end_row=row_n,   end_column=1 + NUM_ODR_YEARS)

col_widths_odr = [2, 10, 8, 7, 7, 7, 6, 6, 8, 7, 10, 10]
for yi in range(NUM_ODR_YEARS):
    bc = yi * ODR_COLS_PER_YEAR + 1
    for di, w in enumerate(col_widths_odr):
        ws_odr.column_dimensions[get_column_letter(bc + di)].width = w
ws_odr.column_dimensions["A"].width = 14
for yi in range(NUM_ODR_YEARS):
    ws_odr.column_dimensions[get_column_letter(2 + yi)].width = 13

print("Sheet 2: ODR ✓")


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3 — MAV SUMMARY (Z-factors all MEVs)
# ─────────────────────────────────────────────────────────────────────────────

ws_mav = wb.create_sheet("MAV Summary")
ws_mav.sheet_view.showGridLines = False

ws_mav.row_dimensions[1].height = 22
sc(ws_mav.cell(1, 1),
   val="MAV Index — Z-Factors Summary (All MEVs, India, IMF WEO Apr 2025)",
   font=FONT_TITLE, align=ALIGN_L)
ws_mav.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=4 + len(DISPLAY_YEARS))
ws_mav.row_dimensions[2].height = 8

ws_mav.row_dimensions[3].height = 16
for ci, lbl in enumerate(["MEV", "LTM", "SD"]):
    sc(ws_mav.cell(3, 1 + ci), val=lbl,
       fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
for ci, yr in enumerate(DISPLAY_YEARS):
    hf = (FILL_DARK if yr <= HIST_CUTOFF else
          FILL_MED  if yr <= 2027 else
          PatternFill("solid", start_color="7F6000", end_color="7F6000"))
    sc(ws_mav.cell(3, 4 + ci), val=str(yr),
       fill=hf, font=FONT_HDR, align=ALIGN_C, border=BORDER)

for ri, code in enumerate(SERIES_ORDER):
    row_e    = 4 + ri
    row_fill = FILL_ALT if ri % 2 == 0 else FILL_WHITE
    ws_mav.row_dimensions[row_e].height = 15
    sc(ws_mav.cell(row_e, 1), val=SERIES_MAP[code],
       fill=row_fill, font=FONT_LABEL, align=ALIGN_L, border=BORDER)
    sc(ws_mav.cell(row_e, 2), val=mav_params[code]["LTM"],
       fill=FILL_GOLD, font=FONT_PARAM, align=ALIGN_R, fmt=FMT_NUM2, border=BORDER)
    sc(ws_mav.cell(row_e, 3), val=mav_params[code]["SD"],
       fill=FILL_GOLD, font=FONT_PARAM, align=ALIGN_R, fmt=FMT_NUM2, border=BORDER)
    for ci, yr in enumerate(DISPLAY_YEARS):
        z = z_factors[code][yr]
        sc(ws_mav.cell(row_e, 4 + ci), val=z,
           fill=year_fill(yr),
           font=FONT_Z_NEG if z < 0 else FONT_Z_POS,
           align=ALIGN_R, fmt=FMT_NUM2, border=BORDER)

ws_mav.row_dimensions[9].height = 8
sc(ws_mav.cell(10, 1),
   val=(f"Z = (Value − LTM) / SD  |  Window {CALIB_YEARS[0]}–{CALIB_YEARS[-1]} "
        f"(n={len(CALIB_YEARS)}, pop. std)  |  "
        "Blue=Actual  Green=IMF Forecast  Amber=Extrapolated"),
   font=FONT_NOTE, align=ALIGN_L)
ws_mav.merge_cells(start_row=10, start_column=1, end_row=10,
                   end_column=4 + len(DISPLAY_YEARS))

ws_mav.column_dimensions["A"].width = 34
ws_mav.column_dimensions["B"].width = 8
ws_mav.column_dimensions["C"].width = 8
for ci in range(len(DISPLAY_YEARS)):
    ws_mav.column_dimensions[get_column_letter(4 + ci)].width = 6.5

print("Sheet 3: MAV Summary ✓")


# ─────────────────────────────────────────────────────────────────────────────
# SHEETS 4-8 — PER-MEV DETAIL
# ─────────────────────────────────────────────────────────────────────────────

for mi, code in enumerate(SERIES_ORDER):
    mev_name = SERIES_MAP[code]
    ws = wb.create_sheet(mev_name[:31])
    ws.sheet_view.showGridLines = False

    ltm = mav_params[code]["LTM"]
    sd  = mav_params[code]["SD"]
    zf  = z_factors[code]

    ws.row_dimensions[1].height = 18
    sc(ws.cell(1, 2), val=mi + 1,
       font=Font(name="Arial", bold=True, size=11, color="1F4E79"), align=ALIGN_C)
    ws.row_dimensions[2].height = 18
    sc(ws.cell(2, 1), val="Selected MEV", font=FONT_LABEL, align=ALIGN_L)
    sc(ws.cell(2, 2), val=mev_name,
       font=Font(name="Arial", bold=True, size=10, color="1F4E79"), align=ALIGN_L)
    ws.row_dimensions[3].height = 8

    ws.row_dimensions[4].height = 16
    sc(ws.cell(4, 1), val="Variable",
       fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
    for ci, yr in enumerate(DISPLAY_YEARS):
        sc(ws.cell(4, 2 + ci), val=str(yr),
           fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
    ws.row_dimensions[5].height = 15
    sc(ws.cell(5, 1), val=mev_name,
       fill=FILL_LIGHT, font=FONT_LABEL, align=ALIGN_L, border=BORDER)
    for ci, yr in enumerate(DISPLAY_YEARS):
        sc(ws.cell(5, 2 + ci), val=round(MAV[code][yr], 2),
           fill=year_fill(yr), font=year_font(yr),
           align=ALIGN_R, fmt=FMT_NUM2, border=BORDER)

    ws.row_dimensions[6].height = 8
    ws.row_dimensions[7].height = 8
    ws.row_dimensions[8].height = 16
    sc(ws.cell(8, 1), val="Scenario Development",
       fill=FILL_MED, font=FONT_SECTION, align=ALIGN_L, border=BORDER)
    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=4)

    ws.row_dimensions[9].height = 15
    sc(ws.cell(9, 1), fill=FILL_MED, border=BORDER)
    sc(ws.cell(9, 2), val=mev_name,
       fill=FILL_MED, font=FONT_HDR, align=ALIGN_C, border=BORDER)
    ws.row_dimensions[10].height = 15
    sc(ws.cell(10, 1), val="Long Term Mean",
       fill=FILL_GOLD, font=FONT_PARAM, align=ALIGN_L, border=BORDER)
    sc(ws.cell(10, 2), val=ltm,
       fill=FILL_GOLD, font=FONT_PARAM, align=ALIGN_R, fmt=FMT_NUM2, border=BORDER)
    ws.row_dimensions[11].height = 15
    sc(ws.cell(11, 1), val="Standard Deviation",
       fill=FILL_GOLD, font=FONT_PARAM, align=ALIGN_L, border=BORDER)
    sc(ws.cell(11, 2), val=sd,
       fill=FILL_GOLD, font=FONT_PARAM, align=ALIGN_R, fmt=FMT_NUM2, border=BORDER)

    ws.row_dimensions[12].height = 8
    ws.row_dimensions[13].height = 8
    ws.row_dimensions[14].height = 16
    sc(ws.cell(14, 1), val="Z_Factor",
       fill=FILL_MED, font=FONT_SECTION, align=ALIGN_L, border=BORDER)
    ws.merge_cells(start_row=14, start_column=1, end_row=14, end_column=4)

    ws.row_dimensions[15].height = 15
    sc(ws.cell(15, 1), val="Variable",
       fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
    sc(ws.cell(15, 2), val="Relationship",
       fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
    for ci, yr in enumerate(DISPLAY_YEARS):
        sc(ws.cell(15, 3 + ci), val=str(yr),
           fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)

    ws.row_dimensions[16].height = 15
    sc(ws.cell(16, 1), val=mev_name,
       fill=FILL_LIGHT, font=FONT_LABEL, align=ALIGN_L, border=BORDER)
    sc(ws.cell(16, 2), fill=FILL_LIGHT, border=BORDER)
    for ci, yr in enumerate(DISPLAY_YEARS):
        z = zf[yr]
        sc(ws.cell(16, 3 + ci), val=z,
           fill=year_fill(yr),
           font=FONT_Z_NEG if z < 0 else FONT_Z_POS,
           align=ALIGN_R, fmt=FMT_NUM2, border=BORDER)

    ws.row_dimensions[17].height = 8
    ws.row_dimensions[18].height = 13
    sc(ws.cell(18, 1),
       val=(f"Z = (Value − LTM) / SD  |  LTM={ltm}, SD={sd}  |  "
            f"Window {CALIB_YEARS[0]}–{CALIB_YEARS[-1]} (n={len(CALIB_YEARS)}, pop. std)  |  "
            "Blue=Actual  Green=IMF Forecast  Amber=Extrapolated"),
       font=FONT_NOTE, align=ALIGN_L)
    ws.merge_cells(start_row=18, start_column=1,
                   end_row=18, end_column=2 + len(DISPLAY_YEARS))

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    for ci in range(len(DISPLAY_YEARS)):
        ws.column_dimensions[get_column_letter(3 + ci)].width = 6.5

print("Sheets 4-8: MEV detail ✓")


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 9 — MEV SCENARIOS
# ─────────────────────────────────────────────────────────────────────────────

ws_sc = wb.create_sheet("MEV Scenarios")
ws_sc.sheet_view.showGridLines = False

ws_sc.row_dimensions[1].height = 22
sc(ws_sc.cell(1, 1),
   val=(f"MEV Scenario — All Variables  "
        f"[GDP-driven  |  LTM={gdp_ltm}  |  SD={gdp_sd}  |  Shock=±{int(SHOCK*100)}%]"),
   font=FONT_TITLE, align=ALIGN_L)
ws_sc.merge_cells(start_row=1, start_column=1, end_row=1,
                  end_column=2 + len(DISPLAY_YEARS))
ws_sc.row_dimensions[2].height = 8

current_row = 3
for mi, code in enumerate(SERIES_ORDER):
    mev_name = SERIES_MAP[code]
    ws_sc.row_dimensions[current_row].height = 14
    sc(ws_sc.cell(current_row, 1), val=f"  {mev_name}",
       fill=FILL_DARK, font=Font(name="Arial", bold=True, color="FFFFFF", size=9),
       align=ALIGN_L, border=BORDER)
    ws_sc.merge_cells(start_row=current_row, start_column=1,
                      end_row=current_row,   end_column=2 + len(DISPLAY_YEARS))
    current_row += 1

    for scen in SCENARIOS:
        ws_sc.row_dimensions[current_row].height = 15
        sc(ws_sc.cell(current_row, 1), val=scen,
           fill=SCEN_HDR_FILL[scen],
           font=Font(name="Arial", bold=True, color="FFFFFF", size=10),
           align=ALIGN_L, border=BORDER)
        ws_sc.merge_cells(start_row=current_row, start_column=1,
                          end_row=current_row,   end_column=2 + len(DISPLAY_YEARS))
        current_row += 1

        ws_sc.row_dimensions[current_row].height = 14
        sc(ws_sc.cell(current_row, 1), val="Variable",
           fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
        for ci, yr in enumerate(DISPLAY_YEARS):
            sc(ws_sc.cell(current_row, 2 + ci), val=str(yr),
               fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
        current_row += 1

        ws_sc.row_dimensions[current_row].height = 14
        sc(ws_sc.cell(current_row, 1), val="MEV_Index",
           fill=SCEN_ROW_FILL[scen], font=FONT_LABEL, align=ALIGN_L, border=BORDER)
        for ci, yr in enumerate(DISPLAY_YEARS):
            z = mev_scenarios[yr][scen]
            sc(ws_sc.cell(current_row, 2 + ci), val=z,
               fill=year_fill(yr), font=SCEN_Z_FONT[scen],
               align=ALIGN_R, fmt=FMT_NUM2, border=BORDER)
        current_row += 1

        for _ in range(2):
            ws_sc.row_dimensions[current_row].height = 5
            current_row += 1

    for _ in range(2):
        ws_sc.row_dimensions[current_row].height = 6
        current_row += 1

ws_sc.row_dimensions[current_row].height = 13
sc(ws_sc.cell(current_row, 1),
   val=(f"Scenario driver: GDP Z-factor  |  "
        f"Base=Z  |  Upturn=Z+|Z|×{SHOCK}  |  Downturn=Z−|Z|×{SHOCK}  |  "
        "Blue=Actual  Green=IMF Forecast  Amber=Extrapolated"),
   font=FONT_NOTE, align=ALIGN_L)
ws_sc.merge_cells(start_row=current_row, start_column=1,
                  end_row=current_row,   end_column=2 + len(DISPLAY_YEARS))

ws_sc.column_dimensions["A"].width = 16
for ci in range(len(DISPLAY_YEARS)):
    ws_sc.column_dimensions[get_column_letter(2 + ci)].width = 6.5

print("Sheet 9: MEV Scenarios ✓")


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 10 — VASICEK PD (stacked scenarios)
# ─────────────────────────────────────────────────────────────────────────────

ws_vasi = wb.create_sheet("Vasicek PD")
ws_vasi.sheet_view.showGridLines = False

BLOCK_ROWS = 3 + len(FROM_BUCKETS) + 2

ws_vasi.row_dimensions[1].height = 20
sc(ws_vasi.cell(1, 1), val="Vasicek Unconditional Probability of Default",
   font=FONT_TITLE, align=ALIGN_L)
ws_vasi.merge_cells(start_row=1, start_column=1, end_row=1,
                    end_column=4 + len(FORECAST_YRS))
ws_vasi.row_dimensions[2].height = 6

for si, scen in enumerate(SCENARIOS):
    base_r   = 3 + si * BLOCK_ROWS
    row_fill = SCEN_ROW_FILL[scen]

    ws_vasi.row_dimensions[base_r].height = 16
    sc(ws_vasi.cell(base_r, 1), val=scen,
       fill=SCEN_HDR_FILL[scen],
       font=Font(name="Arial", bold=True, color="FFFFFF", size=10),
       align=ALIGN_L, border=BORDER)
    ws_vasi.merge_cells(start_row=base_r, start_column=1,
                        end_row=base_r,   end_column=4 + len(FORECAST_YRS))

    ws_vasi.row_dimensions[base_r + 1].height = 15
    for ci, hdr in enumerate(["Grades", "TTC", "Asset Correlation (ρ)"]):
        sc(ws_vasi.cell(base_r + 1, 1 + ci), val=hdr,
           fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
    for ci, yr in enumerate(FORECAST_YRS):
        sc(ws_vasi.cell(base_r + 1, 4 + ci), val=str(yr),
           fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)

    ws_vasi.row_dimensions[base_r + 2].height = 13
    sc(ws_vasi.cell(base_r + 2, 1), val="Z-Factor →",
       fill=FILL_PARAM, font=FONT_PARAM, align=ALIGN_R, border=BORDER)
    sc(ws_vasi.cell(base_r + 2, 2), fill=FILL_PARAM, border=BORDER)
    sc(ws_vasi.cell(base_r + 2, 3), fill=FILL_PARAM, border=BORDER)
    for ci, yr in enumerate(FORECAST_YRS):
        sc(ws_vasi.cell(base_r + 2, 4 + ci), val=scen_z_raw[yr][scen],
           fill=FILL_PARAM, font=FONT_PARAM, align=ALIGN_R,
           fmt=FMT_NUM6, border=BORDER)

    for gi, grade in enumerate(FROM_BUCKETS):
        r = base_r + 3 + gi
        grade_fill = FILL_LIGHT if grade == "90+" else (FILL_ALT if gi % 2 == 0 else FILL_WHITE)
        ws_vasi.row_dimensions[r].height = 15

        sc(ws_vasi.cell(r, 1), val=grade,
           fill=grade_fill,
           font=Font(name="Arial", bold=True, size=9, color="1F4E79"),
           align=ALIGN_C, border=BORDER)
        sc(ws_vasi.cell(r, 2), val=TTC[grade],
           fill=FILL_PARAM, font=FONT_PARAM,
           align=ALIGN_R, fmt=FMT_PCT6, border=BORDER)
        sc(ws_vasi.cell(r, 3), val=RHO[grade],
           fill=FILL_PARAM, font=FONT_PARAM,
           align=ALIGN_R, fmt=FMT_NUM6, border=BORDER)
        for ci, yr in enumerate(FORECAST_YRS):
            sc(ws_vasi.cell(r, 4 + ci), val=pd_results[scen][grade][yr],
               fill=row_fill, font=SCEN_Z_FONT[scen],
               align=ALIGN_R, fmt=FMT_PCT2, border=BORDER)

    for sp in range(2):
        ws_vasi.row_dimensions[base_r + 3 + len(FROM_BUCKETS) + sp].height = 5

note_r = 3 + len(SCENARIOS) * BLOCK_ROWS
for rn, val in enumerate([
    (f"Formula: PD = Φ((Φ⁻¹(TTC) − √ρ·Z) / √(1−ρ))  |  "
     f"Shock=±{int(SHOCK*100)}%  |  LTM={round(gdp_ltm,2)}, SD={round(gdp_sd,2)}"),
    ("TTC=mean ODR per grade  |  ρ=Basel II Retail: 0.03·W+0.16·(1−W)  |  PD format: 0.00%"),
], start=0):
    ws_vasi.row_dimensions[note_r + rn].height = 13
    sc(ws_vasi.cell(note_r + rn, 1), val=val, font=FONT_NOTE, align=ALIGN_L)
    ws_vasi.merge_cells(start_row=note_r+rn, start_column=1,
                        end_row=note_r+rn,   end_column=4+len(FORECAST_YRS))

ws_vasi.column_dimensions["A"].width = 9
ws_vasi.column_dimensions["B"].width = 12
ws_vasi.column_dimensions["C"].width = 22
for ci in range(len(FORECAST_YRS)):
    ws_vasi.column_dimensions[get_column_letter(4 + ci)].width = 8

print("Sheet 10: Vasicek PD ✓")


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 11 — PD SCENARIO COMPARISON
# ─────────────────────────────────────────────────────────────────────────────

ws_cmp = wb.create_sheet("PD Comparison")
ws_cmp.sheet_view.showGridLines = False

ws_cmp.row_dimensions[1].height = 20
sc(ws_cmp.cell(1, 1), val="Vasicek PD — Scenario Comparison",
   font=FONT_TITLE, align=ALIGN_L)
ws_cmp.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=3 + len(SCENARIOS) * len(FORECAST_YRS))
ws_cmp.row_dimensions[2].height = 6

ws_cmp.row_dimensions[3].height = 16
for ci, hdr in enumerate(["Grade", "TTC", "ρ"]):
    sc(ws_cmp.cell(3, 1 + ci), val=hdr,
       fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)
for si, scen in enumerate(SCENARIOS):
    start_col = 4 + si * len(FORECAST_YRS)
    sc(ws_cmp.cell(3, start_col), val=scen,
       fill=SCEN_HDR_FILL[scen],
       font=Font(name="Arial", bold=True, color="FFFFFF", size=9),
       align=ALIGN_C, border=BORDER)
    ws_cmp.merge_cells(start_row=3, start_column=start_col,
                       end_row=3,   end_column=start_col + len(FORECAST_YRS) - 1)

ws_cmp.row_dimensions[4].height = 14
for ci in range(3):
    sc(ws_cmp.cell(4, 1 + ci), fill=FILL_DARK, border=BORDER)
for si, scen in enumerate(SCENARIOS):
    for ci, yr in enumerate(FORECAST_YRS):
        sc(ws_cmp.cell(4, 4 + si * len(FORECAST_YRS) + ci), val=str(yr),
           fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)

for gi, grade in enumerate(FROM_BUCKETS):
    row_e      = 5 + gi
    grade_fill = FILL_LIGHT if grade == "90+" else (FILL_ALT if gi % 2 == 0 else FILL_WHITE)
    ws_cmp.row_dimensions[row_e].height = 15
    sc(ws_cmp.cell(row_e, 1), val=grade,
       fill=grade_fill, font=FONT_LABEL, align=ALIGN_C, border=BORDER)
    sc(ws_cmp.cell(row_e, 2), val=TTC[grade],
       fill=FILL_PARAM, font=FONT_PARAM, align=ALIGN_R, fmt=FMT_PCT6, border=BORDER)
    sc(ws_cmp.cell(row_e, 3), val=RHO[grade],
       fill=FILL_PARAM, font=FONT_PARAM, align=ALIGN_R, fmt=FMT_NUM6, border=BORDER)
    for si, scen in enumerate(SCENARIOS):
        for ci, yr in enumerate(FORECAST_YRS):
            sc(ws_cmp.cell(row_e, 4 + si * len(FORECAST_YRS) + ci),
               val=pd_results[scen][grade][yr],
               fill=SCEN_ROW_FILL[scen], font=SCEN_Z_FONT[scen],
               align=ALIGN_R, fmt=FMT_PCT2, border=BORDER)

ws_cmp.column_dimensions["A"].width = 9
ws_cmp.column_dimensions["B"].width = 12
ws_cmp.column_dimensions["C"].width = 10
for si in range(len(SCENARIOS)):
    for ci in range(len(FORECAST_YRS)):
        ws_cmp.column_dimensions[get_column_letter(4 + si * len(FORECAST_YRS) + ci)].width = 8

print("Sheet 11: PD Comparison ✓")


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 12 — INPUTS AUDIT TRAIL (TTC & ρ)
# ─────────────────────────────────────────────────────────────────────────────

ws_inp = wb.create_sheet("Inputs TTC & rho")
ws_inp.sheet_view.showGridLines = False

ws_inp.row_dimensions[1].height = 20
sc(ws_inp.cell(1, 1),
   val="Model Inputs — TTC (from ODR) and Asset Correlation (Basel II Retail IRB)",
   font=FONT_TITLE, align=ALIGN_L)
ws_inp.merge_cells("A1:H1")
ws_inp.row_dimensions[2].height = 8

ws_inp.row_dimensions[3].height = 15
for ci, hdr in enumerate(["Grade","TTC (precise)","TTC (%)",
                            "ρ (exact)","ρ (rounded)","W (weight)","0.03·W","0.16·(1−W)"]):
    sc(ws_inp.cell(3, 1 + ci), val=hdr,
       fill=FILL_DARK, font=FONT_HDR, align=ALIGN_C, border=BORDER)

for gi, grade in enumerate(FROM_BUCKETS):
    row_e = 4 + gi
    ttc   = TTC[grade]
    rho   = RHO[grade]
    w     = (1 - np.exp(-35 * ttc)) / (1 - np.exp(-35)) if ttc < 1 else 1.0
    rfill = FILL_ALT if gi % 2 == 0 else FILL_WHITE
    ws_inp.row_dimensions[row_e].height = 15
    for ci, (v, fmt) in enumerate(zip(
        [grade, ttc, ttc, rho, round(rho, 2), w, 0.03*w, 0.16*(1-w)],
        ["@", FMT_NUM6, FMT_PCT2, FMT_NUM6, FMT_NUM2, FMT_NUM6, FMT_NUM6, FMT_NUM6]
    )):
        sc(ws_inp.cell(row_e, 1 + ci), val=v, fill=rfill,
           font=FONT_LABEL if ci == 0 else FONT_BODY,
           align=ALIGN_C if ci == 0 else ALIGN_R,
           fmt=fmt, border=BORDER)

ws_inp.row_dimensions[10].height = 8
sc(ws_inp.cell(11, 1),
   val=("TTC = mean ODR per grade (active years only)  |  "
        "ρ = Basel II Retail IRB: 0.03·W + 0.16·(1−W)  |  "
        "W = (1 − e^{−35·TTC}) / (1 − e^{−35})  |  ρ ∈ [0.03, 0.16]"),
   font=FONT_NOTE, align=ALIGN_L)
ws_inp.merge_cells("A11:H11")

for ci, w in enumerate([9, 14, 10, 12, 12, 12, 10, 12]):
    ws_inp.column_dimensions[get_column_letter(1 + ci)].width = w

print("Sheet 12: Inputs TTC & rho ✓")


# ══════════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════════

wb.save(OUTPUT_FILE)
print(f"\n{'='*60}")
print(f"Saved → {OUTPUT_FILE}")
print(f"{'='*60}")
print("Sheets written:")
for i, ws in enumerate(wb.worksheets, 1):
    print(f"  {i:2d}. {ws.title}")